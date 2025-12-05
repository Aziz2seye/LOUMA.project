import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import tempfile
import sys
from pathlib import Path
from PIL import Image
import plotly.express as px
import plotly.graph_objects as go
import os
import plotly.io as pio
from datetime import datetime
import sys
from pathlib import Path

# Configuration pour l'export PNG
pio.kaleido.scope.default_format = "png"
pio.kaleido.scope.default_width = 1200
pio.kaleido.scope.default_height = 600
pio.kaleido.scope.default_scale = 2

# Ajouter le dossier pages au path
current_dir = Path(__file__).parent
if str(current_dir) not in sys.path:
    sys.path.insert(0, str(current_dir))

from db_manager import ReportingDatabase
# Ajouter le répertoire parent au path Python
current_dir = Path(__file__).parent
parent_dir = current_dir.parent
sys.path.insert(0, str(parent_dir))

from utils import load_vto
# 🆕 Import de la base de données
from db_manager import ReportingDatabase

# ====================
# FONCTION D'EXPORT DES GRAPHIQUES EN PNG
# ====================
def download_plotly_as_png(fig, filename):
    """
    Convertit un graphique Plotly en PNG téléchargeable
    """
    import io
    buffer = io.BytesIO()
    fig.write_image(buffer, format='png')
    buffer.seek(0)
    return buffer

# ====================
# Configuration page
# ====================
st.set_page_config(page_title="LOUMA - Reporting", layout="wide", initial_sidebar_state="expanded")

# 🆕 Initialiser la base de données
if 'db_manager' not in st.session_state:
    st.session_state.db_manager = ReportingDatabase()
db = st.session_state.db_manager

# ====================
# CSS personnalisé Orange Sonatel
# ====================
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600;700&display=swap');

    header[data-testid="stHeader"] { display: none; }

    .block-container {
        padding-top: 2rem !important;
        padding-bottom: 2rem !important;
    }

    .main {
        font-family: 'Poppins', sans-serif;
        background: linear-gradient(135deg, #fff5f0 0%, #ffffff 50%, #f0f8ff 100%);
    }

    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #FF7900 0%, #FF5000 100%) !important;
    }

    section[data-testid="stSidebar"] * {
        color: white !important;
    }

    .stDataFrame {
        border-radius: 15px;
        overflow: hidden;
        box-shadow: 0 8px 20px rgba(255, 121, 0, 0.15);
        border: 2px solid #FFE5CC;
    }

    .section-title {
        background: linear-gradient(135deg, #FF7900 0%, #FF5000 100%);
        color: white;
        padding: 0.7rem 1.2rem;
        border-radius: 10px;
        font-weight: 600;
        font-size: 1.1rem;
        margin-bottom: 0.8rem;
        box-shadow: 0 4px 12px rgba(255, 121, 0, 0.25);
        text-align: center;
        max-width: 500px;
        margin-left: auto;
        margin-right: auto;
    }

    .stButton > button {
        background: linear-gradient(135deg, #FF7900 0%, #FF5000 100%);
        color: white;
        border: none;
        border-radius: 10px;
        padding: 0.8rem 2rem;
        font-weight: 600;
        font-size: 1.1rem;
        box-shadow: 0 4px 12px rgba(255, 121, 0, 0.3);
        transition: all 0.3s ease;
        width: 100%;
    }

    .stButton > button:hover {
        background: linear-gradient(135deg, #FF5000 0%, #FF3000 100%);
        box-shadow: 0 6px 18px rgba(255, 121, 0, 0.5);
        transform: translateY(-2px);
    }

    .stDownloadButton > button {
        background: linear-gradient(135deg, #00D4AA 0%, #00B890 100%);
        color: white;
        border: none;
        border-radius: 10px;
        padding: 0.8rem 2rem;
        font-weight: 600;
        font-size: 1.1rem;
        box-shadow: 0 4px 12px rgba(0, 212, 170, 0.3);
        transition: all 0.3s ease;
        width: 100%;
    }

    .stDownloadButton > button:hover {
        background: linear-gradient(135deg, #00B890 0%, #009A7A 100%);
        box-shadow: 0 6px 18px rgba(0, 212, 170, 0.5);
        transform: translateY(-2px);
    }

    .metric-card {
        background: white;
        border-radius: 12px;
        padding: 1.5rem;
        box-shadow: 0 4px 15px rgba(255, 121, 0, 0.15);
        border: 2px solid #FFE5CC;
        text-align: center;
    }

    .metric-value {
        font-size: 2rem;
        font-weight: 700;
        color: #FF7900;
    }

    .metric-label {
        font-size: 1rem;
        color: #666;
        margin-top: 0.5rem;
    }

    .info-card {
        background: linear-gradient(135deg, #009CA6 0%, #00B8C5 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 12px;
        box-shadow: 0 4px 15px rgba(0, 156, 166, 0.3);
        margin-bottom: 1.5rem;
    }

    @keyframes fadeIn {
        from { opacity: 0; transform: translateY(20px); }
        to { opacity: 1; transform: translateY(0); }
    }

    .stDataFrame, .metric-card, .info-card {
        animation: fadeIn 0.5s ease-out;
    }
</style>
""", unsafe_allow_html=True)

# ====================
# Charger le logo
# ====================
logo = None
logo_paths = [
    parent_dir / "assets" / "logo sonatel.png",
    Path("assets") / "logo sonatel.png",
    Path("../assets/logo sonatel.png"),
]

for logo_path in logo_paths:
    try:
        if logo_path.exists():
            logo = Image.open(logo_path)
            break
    except:
        continue

# ====================
# Header avec logo et titre
# ====================
col_logo, col_title = st.columns([1, 3])

with col_logo:
    if logo:
        st.image(logo, width=280)

with col_title:
    st.markdown("""
    <div style="
        background: linear-gradient(135deg, #FF7900 0%, #FF5000 100%);
        padding: 2rem;
        border-radius: 20px;
        box-shadow: 0 8px 25px rgba(255, 121, 0, 0.4);
        display: flex;
        flex-direction: column;
        justify-content: center;
        border: 3px solid rgba(255, 255, 255, 0.2);
        height: 100%;
    ">
        <h1 style="
            color: white;
            font-size: 2.5rem;
            font-weight: 700;
            margin: 0;
            text-shadow: 3px 3px 10px rgba(0, 0, 0, 0.3);
        ">
            📈 Générateur de Reporting
        </h1>
        <p style="
            color: rgba(255, 255, 255, 0.95);
            font-size: 1.2rem;
            margin: 0.8rem 0 0 0;
            font-weight: 400;
        ">
            Reporting Journalier & Hebdomadaire - Orange Sénégal
        </p>
    </div>
    """, unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# 🎯 Mapping DRV unique
DRV_MAPPING = {
    "DV-DRV2_DIRECTION REGIONALE DES VENTES DAKAR 2": "DR2",
    "DV-DRVS_DIRECTION REGIONALE DES VENTES SUD": "DRS",
    "DV-DRVSE_DIRECTION REGIONALE DES VENTES SUD-EST": "DRSE",
    "DV-DRVN_DIRECTION REGIONALE DES VENTES NORD": "DRN",
    "DV-DRVC_DIRECTION REGIONALE DES VENTES CENTRE": "DRC",
    "DV-DRVE_DIRECTION REGIONALE DES VENTES EST": "DRE"
}

# 🔁 Bouton pour revenir à la sélection
if st.session_state.get("reporting_type"):
    if st.button("↩ Retour au menu principal"):
        st.session_state.reporting_type = None
        st.rerun()

if not st.session_state.get("reporting_type"):
    st.markdown('<div class="section-title">Choisissez un type de reporting</div>', unsafe_allow_html=True)

    st.markdown("""
    <div class="info-card">
        <h3 style="margin: 0 0 0.5rem 0;">ℹ Information</h3>
        <p style="margin: 0;">Sélectionnez le type de rapport : journalier pour les performances quotidiennes ou hebdomadaire pour un résumé de la semaine.</p>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)
    with col1:
        if st.button("🕐 Reporting Journalier", use_container_width=True):
            st.session_state.reporting_type = "journalier"
            st.rerun()
    with col2:
        if st.button("📅 Reporting Hebdomadaire", use_container_width=True):
            st.session_state.reporting_type = "hebdomadaire"
            st.rerun()
    with col3:
        if st.button("📊 Historique & Stats", use_container_width=True):
            st.session_state.reporting_type = "historique"
            st.rerun()


# 🚀 Bloc principal : Reporting Journalier
if st.session_state.get("reporting_type") == "journalier":
    st.markdown('<div class="section-title">🕐 Reporting Journalier</div>', unsafe_allow_html=True)

    uploaded_file = st.file_uploader("📁 Importer le fichier Excel brut (Journalier)", type=["xlsx", "csv"])

    if uploaded_file:
        st.markdown("<br>", unsafe_allow_html=True)

        if uploaded_file.name.endswith('.csv'):
            df = pd.read_csv(uploaded_file, encoding='utf-8', sep='|')
        else:
            xls = pd.ExcelFile(uploaded_file)
            sheet_names = xls.sheet_names
            selected_sheet = st.selectbox("🗂 Choisir la feuille à exploiter :", options=sheet_names)
            df = pd.read_excel(uploaded_file, sheet_name=selected_sheet)

        st.markdown("<br>", unsafe_allow_html=True)

        # ✅ Charger logins depuis fichier VTO
        vto_df = load_vto()
        logins_concernes = vto_df["LOGIN"].astype(str).str.lower().tolist()
        details = ["En Cours-Identification", "Identifie", "Identifie Photo"]

        # ✅ Mapping des colonnes
        column_mapping = {
            'MSISDN': 'TOTAL_SIM',
            'ACCUEIL_VENDEUR': 'PVT',
            'LOGIN_VENDEUR': 'LOGIN',
            'AGENCE_VENDEUR': 'DR'
        }

        missing_columns = [col for col in column_mapping.keys() if col not in df.columns]

        if missing_columns:
            st.error(f"❌ Colonnes manquantes : {', '.join(missing_columns)}")
            st.stop()

        df = df.rename(columns=column_mapping)

        if 'LOGIN' not in df.columns:
            st.error("❌ La colonne LOGIN n'a pas pu être créée.")
            st.stop()

        # Nettoyage des données
        df['LOGIN'] = df['LOGIN'].astype(str).str.lower()
        df['DR'] = df['DR'].astype(str).str.strip().str.upper()
        df['NOM_VENDEUR'] = df['NOM_VENDEUR'].astype(str).str.strip().str.upper()
        df['PRENOM_VENDEUR'] = df['PRENOM_VENDEUR'].astype(str).str.strip().str.upper()

        # 🔍 Filtrage
        df_filtre = df[df['LOGIN'].isin(logins_concernes) & df['ETAT_IDENTIFICATION'].astype(str).isin(details)]
        df_filtre["DR"] = df_filtre["DR"].replace(DRV_MAPPING)

        st.success(f"✅ Fichier filtré avec succès ! {df_filtre.shape[0]} ventes journalières")

        # Métriques
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-value">{df_filtre.shape[0]}</div>
                <div class="metric-label">📊 Ventes du Jour</div>
            </div>
            """, unsafe_allow_html=True)
        with col2:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-value">{df_filtre['LOGIN'].nunique()}</div>
                <div class="metric-label">👤 VTO Actifs</div>
            </div>
            """, unsafe_allow_html=True)
        with col3:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-value">{df_filtre['PVT'].nunique()}</div>
                <div class="metric-label">🏪 PVT Concernés</div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # 📊 TABLEAU 1 : Résumé par PVT
        st.markdown('<div class="section-title">📊 Résumé par PVT</div>', unsafe_allow_html=True)

        df_pvt_summary = df_filtre.groupby(['DR', 'PVT'], as_index=False).size()
        df_pvt_summary.columns = ['DR', 'PVT', 'TOTAL_SIM']
        df_pvt_summary['OBJECTIF'] = 240  # Objectif journalier par PVT
        df_pvt_summary['TR'] = (df_pvt_summary['TOTAL_SIM'] / df_pvt_summary['OBJECTIF'] * 100).round(0).astype(int).astype(str) + '%'

        # Trier par DRV
        df_pvt_summary = df_pvt_summary.sort_values(['DR', 'PVT'])

        # Ajouter ligne de total pour l'affichage
        total_sim = df_pvt_summary['TOTAL_SIM'].sum()
        total_objectif = df_pvt_summary['OBJECTIF'].sum()
        total_tr = round((total_sim / total_objectif * 100), 1)

        df_pvt_summary_display = df_pvt_summary.copy()
        df_pvt_summary_display.loc[len(df_pvt_summary_display)] = ['', 'TOTAL', total_sim, total_objectif, f'{total_tr}%']

        st.dataframe(df_pvt_summary_display, use_container_width=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # 📊 TABLEAU 2 : Détails par VTO
        st.markdown('<div class="section-title">📋 Détails par VTO</div>', unsafe_allow_html=True)

        df_reporting = df_filtre.groupby(['DR', 'PVT', 'PRENOM_VENDEUR', 'NOM_VENDEUR', 'LOGIN'], as_index=False).size()
        df_reporting.columns = ['DR', 'PVT', 'PRENOM_VENDEUR', 'NOM_VENDEUR', 'LOGIN', 'TOTAL_SIM']

        # Trier par DRV, PVT puis TOTAL_SIM
        df_reporting = df_reporting.sort_values(['DR', 'PVT', 'TOTAL_SIM'], ascending=[True, True, False])

        # Ajouter ligne de total pour l'affichage
        total_sim_vto = df_reporting['TOTAL_SIM'].sum()
        df_reporting_display = df_reporting.copy()
        df_reporting_display.loc[len(df_reporting_display)] = ['', '', '', '', 'TOTAL', total_sim_vto]

        st.dataframe(df_reporting_display, use_container_width=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # 🆕 SECTION SAUVEGARDE DANS LA BASE DE DONNÉES
        st.markdown('<div class="section-title">💾 Sauvegarde dans la Base de Données</div>', unsafe_allow_html=True)

        col_date, col_save, col_load = st.columns([2, 1, 1])

        with col_date:
            date_reporting = st.date_input(
                "📅 Date du reporting",
                value=datetime.now().date(),
                format="DD/MM/YYYY",
                key="date_jour"
            )
            date_str = date_reporting.strftime('%Y-%m-%d')

        with col_save:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("💾 Sauvegarder", use_container_width=True, key="save_jour"):
                # Supprimer les lignes TOTAL avant sauvegarde
                df_pvt_save = df_pvt_summary.copy()
                df_vto_save = df_reporting.copy()

                success, message = db.save_daily_report(
                    date_str,
                    df_pvt_save,
                    df_vto_save
                )
                if success:
                    st.success(f"✅ {message}")
                else:
                    st.error(f"❌ {message}")

        with col_load:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("📂 Charger", use_container_width=True, key="show_load_jour"):
                st.session_state.show_load_journalier = True

        # Afficher le sélecteur de chargement si demandé
        if st.session_state.get('show_load_journalier'):
            st.markdown("---")
            available_dates = db.get_available_dates('daily')

            if available_dates:
                col_select, col_action, col_cancel = st.columns([3, 1, 1])

                with col_select:
                    date_options = [d[0] for d in available_dates]
                    selected_date = st.selectbox(
                        "Choisir une date à charger",
                        options=date_options,
                        format_func=lambda x: datetime.strptime(x, '%Y-%m-%d').strftime('%d/%m/%Y'),
                        key="select_date_jour"
                    )

                with col_action:
                    st.markdown("<br>", unsafe_allow_html=True)
                    if st.button("📥 Charger", key="confirm_load_jour"):
                        st.info(f"📊 Données du {datetime.strptime(selected_date, '%Y-%m-%d').strftime('%d/%m/%Y')} chargées depuis la BDD")
                        st.session_state.show_load_journalier = False
                        st.rerun()

                with col_cancel:
                    st.markdown("<br>", unsafe_allow_html=True)
                    if st.button("❌ Annuler", key="cancel_load_jour"):
                        st.session_state.show_load_journalier = False
                        st.rerun()
            else:
                st.info("Aucun reporting sauvegardé dans la base de données")
                if st.button("Fermer"):
                    st.session_state.show_load_journalier = False
                    st.rerun()

        st.markdown("<br>", unsafe_allow_html=True)

        # 📊 GRAPHIQUES - VERSION AMÉLIORÉE

        st.markdown('<div class="section-title">📊 VISUALISATION DES PERFORMANCES</div>', unsafe_allow_html=True)

        # Première ligne : Top DR (circulaire + barres VERTICALES)
        st.markdown('<h4 style="text-align: center; color: #FF7900;">🗺 Distribution par Direction Régionale</h4>', unsafe_allow_html=True)

        df_drv = df_reporting.groupby('DR').agg({'TOTAL_SIM': 'sum'}).reset_index()
        df_drv = df_drv.sort_values('TOTAL_SIM', ascending=False)  # Tri décroissant

        # Calculer le pourcentage pour chaque DR
        total_ventes = df_drv['TOTAL_SIM'].sum()
        df_drv['POURCENTAGE'] = (df_drv['TOTAL_SIM'] / total_ventes * 100).round(1)

        col_dr1, col_dr2 = st.columns(2)

        with col_dr1:
            # Diagramme circulaire
            fig_pie_dr = px.pie(
                df_drv,
                values='TOTAL_SIM',
                names='DR',
                title='Distribution des ventes par DR (Journalier)',
                color_discrete_sequence=['#FF7900', '#FF5000', '#FF3000', '#E57200', '#CC6600', '#B35900'],
                hole=0.3
            )

            df_drv['LABEL'] = df_drv.apply(lambda row: f"{row['DR']}<br>{row['TOTAL_SIM']} ventes<br>({row['POURCENTAGE']}%)", axis=1)

            fig_pie_dr.update_traces(
                textposition='inside',
                textinfo='text',
                text=df_drv['LABEL'],
                textfont_size=10,
                marker=dict(line=dict(color='white', width=2)),
                hovertemplate="<b>%{label}</b><br>" +
                             "Ventes: %{value}<br>" +
                             "Pourcentage: %{percent}<br>" +
                             "<extra></extra>"
            )

            fig_pie_dr.update_layout(
                title=dict(
                    text="Distribution des ventes par DR (Journalier)",
                    font=dict(size=16, family='Poppins', color='#FF7900'),
                    x=0.5,
                    y=0.95
                ),
                font=dict(family='Poppins', size=10),
                height=400,
                margin=dict(t=70, b=30, l=30, r=30),
                showlegend=False,
                hoverlabel=dict(
                    bgcolor="white",
                    font_size=11,
                    font_family="Poppins"
                )
            )

            st.plotly_chart(fig_pie_dr, use_container_width=True)

            # Bouton de téléchargement
            st.download_button(
                label="📥 Télécharger (PNG)",
                data=download_plotly_as_png(fig_pie_dr, "Distribution_DR_Circulaire"),
                file_name="Distribution_DR_Circulaire.png",
                mime="image/png",
                key="download_dr_pie"
            )

        with col_dr2:
            # Diagramme à barres VERTICALES triées
            fig_bar_dr = go.Figure()

            # Tri pour avoir les plus hautes barres à gauche
            df_drv_bar = df_drv.sort_values('TOTAL_SIM', ascending=False)

            fig_bar_dr.add_trace(go.Bar(
                x=df_drv_bar['DR'],
                y=df_drv_bar['TOTAL_SIM'],
                marker_color='#FF7900',
                text=df_drv_bar.apply(lambda row: f"{row['TOTAL_SIM']} ventes", axis=1),
                textposition='outside',
                textfont=dict(size=10, color='#333'),
                textangle=0,
                hovertemplate="<b>%{x}</b><br>" +
                             "Ventes: %{y}<br>" +
                             "Pourcentage: %{customdata}%<br>" +
                             "<extra></extra>",
                customdata=df_drv_bar['POURCENTAGE']
            ))

            fig_bar_dr.update_layout(
                title=dict(
                    text="Performances par DR (Journalier)",
                    font=dict(size=16, family='Poppins', color='#FF7900'),
                    x=0.5,
                    y=0.95
                ),
                xaxis_title='',
                yaxis_title='Nombre de ventes',
                template='plotly_white',
                height=400,
                margin=dict(t=70, b=30, l=60, r=30),
                font=dict(family='Poppins', size=10),
                hoverlabel=dict(
                    bgcolor="white",
                    font_size=11,
                    font_family="Poppins"
                ),
                xaxis=dict(
                    tickangle=0,
                    tickfont=dict(size=10),
                    automargin=True,
                    showgrid=False,
                    categoryorder='total descending'  # Tri décroissant
                ),
                yaxis=dict(
                    showgrid=True,
                    gridcolor='lightgray'
                )
            )

            st.plotly_chart(fig_bar_dr, use_container_width=True)

            # Bouton de téléchargement
            st.download_button(
                label="📥 Télécharger (PNG)",
                data=download_plotly_as_png(fig_bar_dr, "Distribution_DR_Barres"),
                file_name="Distribution_DR_Barres.png",
                mime="image/png",
                key="download_dr_bar"
            )

        # Deuxième ligne : Top PVT (Top 5) et Top VTO (Top 10)
        col_pvt, col_vto = st.columns(2)

        with col_pvt:
            # TOP 5 PVT (Diagramme horizontal)
            st.markdown('<h4 style="text-align: center; color: #009CA6;">🏪 Top 5 Points de Vente</h4>', unsafe_allow_html=True)

            # Calculer le top 5 PVT
            df_pvt_summary_chart = df_filtre.groupby(['DR', 'PVT']).agg({'TOTAL_SIM': 'sum'}).reset_index()
            df_top_pvt = df_pvt_summary_chart.nlargest(5, 'TOTAL_SIM')  # Top 5 seulement
            df_top_pvt = df_top_pvt.sort_values('TOTAL_SIM', ascending=True)  # Tri pour meilleur en haut

            # Créer un label court
            def create_pvt_label(pvt_name, dr, max_length=25):
                # Raccourcir le nom si trop long
                short_name = pvt_name[:max_length-3] + "..." if len(pvt_name) > max_length else pvt_name
                return f"{short_name} (DR: {dr})"

            df_top_pvt['LABEL'] = df_top_pvt.apply(
                lambda row: create_pvt_label(row['PVT'], row['DR']),
                axis=1
            )

            fig_pvt = go.Figure()

            # Barres horizontales - SUPPRIMER LE TEXTE À DROITE
            fig_pvt.add_trace(go.Bar(
                y=df_top_pvt['LABEL'],
                x=df_top_pvt['TOTAL_SIM'],
                orientation='h',
                marker_color='#009CA6',
                # SUPPRIMÉ: text=df_top_pvt['TOTAL_SIM'],
                # SUPPRIMÉ: textposition='outside',
                # SUPPRIMÉ: textfont=dict(size=10, color='#333'),
                # SUPPRIMÉ: textangle=0,
                hovertemplate="<b>%{customdata[0]}</b><br>" +
                             "Ventes: %{x}<br>" +
                             "DR: %{customdata[1]}<br>" +
                             "<extra></extra>",
                customdata=df_top_pvt[['PVT', 'DR']]
            ))

            fig_pvt.update_layout(
                title=dict(
                    text="Top 5 PVT - Performances Journalières",
                    font=dict(size=16, family='Poppins', color='#009CA6'),
                    x=0.5,
                    y=0.95
                ),
                xaxis_title='Nombre de ventes',
                yaxis_title='',
                template='plotly_white',
                height=350,  # Hauteur réduite pour 5 éléments
                margin=dict(t=70, b=30, l=250, r=30),  # Grande marge gauche pour les longs labels
                font=dict(family='Poppins', size=10),
                hoverlabel=dict(
                    bgcolor="white",
                    font_size=11,
                    font_family="Poppins"
                ),
                yaxis=dict(
                    tickfont=dict(size=9),
                    automargin=True,
                    showgrid=False
                ),
                xaxis=dict(
                    showgrid=True,
                    gridcolor='lightgray'
                )
            )

            st.plotly_chart(fig_pvt, use_container_width=True)

            # Bouton de téléchargement
            st.download_button(
                label="📥 Télécharger (PNG)",
                data=download_plotly_as_png(fig_pvt, "Top_5_PVT"),
                file_name="Top_5_PVT.png",
                mime="image/png",
                key="download_pvt_chart"
            )

        with col_vto:
            # TOP 10 VTO (Diagramme vertical)
            st.markdown('<h4 style="text-align: center; color: #FF5000;">👥 Top 10 Vendeurs (VTO)</h4>', unsafe_allow_html=True)

            # Top 10 VTO
            df_top10 = df_reporting.nlargest(10, 'TOTAL_SIM').copy()
            df_top10 = df_top10.sort_values('TOTAL_SIM', ascending=False)

            # Créer un label complet
            def create_vto_label(prenom, nom, pvt, dr, max_length=20):
                nom_complet = f"{prenom} {nom}"
                # Raccourcir le nom complet si trop long
                if len(nom_complet) > max_length:
                    nom_complet = nom_complet[:max_length-3] + "..."

                return f"{nom_complet}<br>(PVT: {pvt[:12]}...)" if len(pvt) > 12 else f"{nom_complet}<br>(PVT: {pvt})"

            df_top10['LABEL'] = df_top10.apply(
                lambda row: create_vto_label(
                    row['PRENOM_VENDEUR'],
                    row['NOM_VENDEUR'],
                    row['PVT'],
                    row['DR']
                ),
                axis=1
            )

            fig_top10 = go.Figure()

            fig_top10.add_trace(go.Bar(
                x=df_top10['LABEL'],
                y=df_top10['TOTAL_SIM'],
                marker_color='#FF5000',
                text=df_top10['TOTAL_SIM'],
                textposition='outside',
                textfont=dict(size=10, color='#333'),
                marker_line=dict(color='white', width=1),
                hovertemplate="<b>%{customdata[0]} %{customdata[1]}</b><br>" +
                             "Ventes: %{y}<br>" +
                             "PVT: %{customdata[2]}<br>" +
                             "DR: %{customdata[3]}<br>" +
                             "<extra></extra>",
                customdata=df_top10[['PRENOM_VENDEUR', 'NOM_VENDEUR', 'PVT', 'DR']]
            ))

            fig_top10.update_layout(
                title=dict(
                    text="Top 10 Vendeurs (VTO) - Journalier",
                    font=dict(size=16, family='Poppins', color='#FF5000'),
                    x=0.5,
                    y=0.95
                ),
                xaxis_title='',
                yaxis_title='Nombre de ventes',
                template='plotly_white',
                height=400,
                margin=dict(t=70, b=100, l=60, r=30),
                font=dict(family='Poppins', size=10),
                hoverlabel=dict(
                    bgcolor="white",
                    font_size=11,
                    font_family="Poppins"
                ),
                xaxis=dict(
                    tickangle=-35,
                    tickfont=dict(size=9),
                    automargin=True,
                    showgrid=False
                ),
                yaxis=dict(
                    showgrid=True,
                    gridcolor='lightgray'
                )
            )

            st.plotly_chart(fig_top10, use_container_width=True)

            # Bouton de téléchargement
            st.download_button(
                label="📥 Télécharger ce graphique (PNG pour PowerPoint)",
                data=download_plotly_as_png(fig_top10, "Top_VTO"),
                file_name="Top_VTO.png",
                mime="image/png",
                key="download_top10"
            )

        # 🔹 MÉTRIQUES DE PERFORMANCE SIMPLIFIÉES
        st.markdown('<br><br>', unsafe_allow_html=True)
        st.markdown('<div class="section-title">📈 RÉSUMÉ DES PERFORMANCES</div>', unsafe_allow_html=True)

        col1, col2, col3, col4 = st.columns(4)

        with col1:
            st.markdown(f"""
            <div style="
                background: white;
                border-radius: 12px;
                padding: 1.2rem;
                box-shadow: 0 4px 15px rgba(255, 121, 0, 0.1);
                border-left: 5px solid #FF7900;
                text-align: center;
            ">
                <div style="font-size: 2rem; font-weight: 700; color: #FF7900;">{df_filtre.shape[0]}</div>
                <div style="font-size: 0.9rem; color: #666; margin-top: 0.5rem;">📊 Total Ventes</div>
            </div>
            """, unsafe_allow_html=True)

        with col2:
            meilleur_dr = df_drv.nlargest(1, 'TOTAL_SIM')
            meilleur_dr_nom = meilleur_dr['DR'].iloc[0] if not meilleur_dr.empty else "N/A"
            meilleur_dr_ventes = meilleur_dr['TOTAL_SIM'].iloc[0] if not meilleur_dr.empty else 0

            st.markdown(f"""
            <div style="
                background: white;
                border-radius: 12px;
                padding: 1.2rem;
                box-shadow: 0 4px 15px rgba(0, 156, 166, 0.1);
                border-left: 5px solid #009CA6;
                text-align: center;
            ">
                <div style="font-size: 1.8rem; font-weight: 700; color: #009CA6;">{meilleur_dr_nom}</div>
                <div style="font-size: 0.9rem; color: #666; margin-top: 0.5rem;">🏆 Meilleure DR</div>
            </div>
            """, unsafe_allow_html=True)

        with col3:
            meilleur_pvt = df_top_pvt.nlargest(1, 'TOTAL_SIM')
            meilleur_pvt_nom = meilleur_pvt['PVT'].iloc[0][:15] + "..." if not meilleur_pvt.empty and len(meilleur_pvt['PVT'].iloc[0]) > 15 else (meilleur_pvt['PVT'].iloc[0] if not meilleur_pvt.empty else "N/A")
            meilleur_pvt_ventes = meilleur_pvt['TOTAL_SIM'].iloc[0] if not meilleur_pvt.empty else 0

            st.markdown(f"""
            <div style="
                background: white;
                border-radius: 12px;
                padding: 1.2rem;
                box-shadow: 0 4px 15px rgba(0, 212, 170, 0.1);
                border-left: 5px solid #00D4AA;
                text-align: center;
            ">
                <div style="font-size: 1.8rem; font-weight: 700; color: #00D4AA;">{meilleur_pvt_nom}</div>
                <div style="font-size: 0.9rem; color: #666; margin-top: 0.5rem;">🏪 Meilleur PVT</div>
            </div>
            """, unsafe_allow_html=True)

        with col4:
            meilleur_vto = df_top10.nlargest(1, 'TOTAL_SIM')
            meilleur_vto_nom = meilleur_vto['PRENOM_VENDEUR'].iloc[0][:8] + "..." if not meilleur_vto.empty and len(meilleur_vto['PRENOM_VENDEUR'].iloc[0]) > 8 else (meilleur_vto['PRENOM_VENDEUR'].iloc[0] if not meilleur_vto.empty else "N/A")
            meilleur_vto_ventes = meilleur_vto['TOTAL_SIM'].iloc[0] if not meilleur_vto.empty else 0

            st.markdown(f"""
            <div style="
                background: white;
                border-radius: 12px;
                padding: 1.2rem;
                box-shadow: 0 4px 15px rgba(255, 80, 0, 0.1);
                border-left: 5px solid #FF5000;
                text-align: center;
            ">
                <div style="font-size: 1.8rem; font-weight: 700; color: #FF5000;">{meilleur_vto_nom}</div>
                <div style="font-size: 0.9rem; color: #666; margin-top: 0.5rem;">👤 Meilleur VTO</div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<div class="section-title">📥 Téléchargement du Rapport Excel</div>', unsafe_allow_html=True)

        # 🧾 Export Excel AVEC LES DEUX TABLEAUX
        try:
            # Créer les DataFrames avec totaux
            total_sim = int(df_pvt_summary['TOTAL_SIM'].sum())
            total_objectif = int(df_pvt_summary['OBJECTIF'].sum())
            total_tr = round((total_sim / total_objectif * 100), 1)

            df_pvt_summary_export = df_pvt_summary.copy().reset_index(drop=True)
            new_row_pvt = {
                'DR': '',
                'PVT': 'TOTAL',
                'TOTAL_SIM': total_sim,
                'OBJECTIF': total_objectif,
                'TR': f'{total_tr}%'
            }
            df_pvt_summary_export = pd.concat([
                df_pvt_summary_export,
                pd.DataFrame([new_row_pvt])
            ], ignore_index=True)

            total_sim_vto = int(df_reporting['TOTAL_SIM'].sum())
            df_reporting_export = df_reporting.copy().reset_index(drop=True)
            new_row_vto = {
                'DR': '',
                'PVT': '',
                'PRENOM_VENDEUR': '',
                'NOM_VENDEUR': '',
                'LOGIN': 'TOTAL',
                'TOTAL_SIM': total_sim_vto
            }
            df_reporting_export = pd.concat([
                df_reporting_export,
                pd.DataFrame([new_row_vto])
            ], ignore_index=True)

            buffer_output = BytesIO()

            with pd.ExcelWriter(buffer_output, engine='openpyxl') as writer:
                df_pvt_summary_export.to_excel(writer, sheet_name='Résumé PVT', index=False)
                df_reporting_export.to_excel(writer, sheet_name='Détails VTO', index=False)

            buffer_output.seek(0)
            wb = load_workbook(buffer_output)

            # Formater la feuille "Résumé PVT"
            ws_pvt = wb['Résumé PVT']

            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            header_font = Font(bold=True, size=11)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for cell in ws_pvt[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            drv_ranges_pvt = []
            current_drv = None
            drv_start = 2

            for row_idx in range(2, ws_pvt.max_row):
                drv_value = ws_pvt.cell(row_idx, 1).value
                if drv_value and drv_value != current_drv:
                    if current_drv is not None and row_idx > drv_start:
                        drv_ranges_pvt.append((drv_start, row_idx - 1, current_drv))
                    current_drv = drv_value
                    drv_start = row_idx

            if ws_pvt.max_row - 1 >= drv_start:
                drv_ranges_pvt.append((drv_start, ws_pvt.max_row - 1, current_drv))

            for start_row, end_row, drv_value in drv_ranges_pvt:
                if end_row > start_row:
                    ws_pvt.merge_cells(f'A{start_row}:A{end_row}')
                    ws_pvt.cell(start_row, 1).alignment = Alignment(horizontal='left', vertical='center')
                    ws_pvt.cell(start_row, 1).font = Font(bold=True, size=10)

            ws_pvt.merge_cells(f'A{ws_pvt.max_row}:B{ws_pvt.max_row}')
            ws_pvt.cell(ws_pvt.max_row, 1).value = 'TOTAL'

            for row_idx in range(2, ws_pvt.max_row + 1):
                for col_idx in range(1, 6):
                    cell = ws_pvt.cell(row_idx, col_idx)
                    cell.border = thin_border

                    if row_idx == ws_pvt.max_row:
                        cell.font = Font(bold=True, size=11)
                        cell.fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
                        if col_idx in [1, 3, 4, 5]:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif col_idx in [1, 2]:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        cell.font = Font(size=10)
                    elif col_idx in [3, 4]:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(size=10)
                    elif col_idx == 5:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(bold=True, size=10)

            ws_pvt.column_dimensions['A'].width = 8
            ws_pvt.column_dimensions['B'].width = 45
            ws_pvt.column_dimensions['C'].width = 12
            ws_pvt.column_dimensions['D'].width = 12
            ws_pvt.column_dimensions['E'].width = 10
            ws_pvt.freeze_panes = 'A2'

            # Formater la feuille "Détails VTO"
            ws = wb['Détails VTO']

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            drv_ranges = []
            current_drv = None
            drv_start = 2

            for row_idx in range(2, ws.max_row):
                drv_value = ws.cell(row_idx, 1).value
                if drv_value != current_drv:
                    if current_drv is not None and row_idx > drv_start:
                        drv_ranges.append((drv_start, row_idx - 1, current_drv))
                    current_drv = drv_value
                    drv_start = row_idx

            if ws.max_row - 1 >= drv_start:
                drv_ranges.append((drv_start, ws.max_row - 1, current_drv))

            pvt_ranges = []
            current_pvt = None
            pvt_start = 2

            for row_idx in range(2, ws.max_row):
                pvt_value = ws.cell(row_idx, 2).value
                if pvt_value != current_pvt:
                    if current_pvt is not None and row_idx > pvt_start:
                        pvt_ranges.append((pvt_start, row_idx - 1, current_pvt))
                    current_pvt = pvt_value
                    pvt_start = row_idx

            if ws.max_row - 1 >= pvt_start:
                pvt_ranges.append((pvt_start, ws.max_row - 1, current_pvt))

            for start_row, end_row, drv_value in drv_ranges:
                if end_row > start_row:
                    ws.merge_cells(f'A{start_row}:A{end_row}')
                    ws.cell(start_row, 1).alignment = Alignment(horizontal='left', vertical='center')
                    ws.cell(start_row, 1).font = Font(bold=True, size=10)

            for start_row, end_row, pvt_value in pvt_ranges:
                if end_row > start_row:
                    ws.merge_cells(f'B{start_row}:B{end_row}')
                    ws.cell(start_row, 2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws.cell(start_row, 2).font = Font(bold=True, size=10)

            ws.merge_cells(f'A{ws.max_row}:E{ws.max_row}')
            ws.cell(ws.max_row, 1).value = 'TOTAL'

            for row_idx in range(2, ws.max_row + 1):
                for col_idx in range(1, 7):
                    cell = ws.cell(row_idx, col_idx)
                    cell.border = thin_border

                    if row_idx == ws.max_row:
                        cell.font = Font(bold=True, size=11)
                        cell.fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
                        if col_idx in [1, 6]:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif col_idx in [1, 2, 3, 4, 5]:
                        if col_idx not in [1, 2]:
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                            cell.font = Font(size=10)
                    elif col_idx == 6:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(size=10)

            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 45
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 12
            ws.freeze_panes = 'A2'

            final_buffer = BytesIO()
            wb.save(final_buffer)
            final_buffer.seek(0)

            st.download_button(
                label="📥 Télécharger le Reporting Journalier (2 feuilles : Résumé PVT + Détails VTO)",
                data=final_buffer,
                file_name="Daily_Reporting.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"❌ Erreur lors de la génération du fichier Excel : {str(e)}")
            import traceback
            st.code(traceback.format_exc())

# 🚀 Bloc principal : Reporting Hebdomadaire
if st.session_state.get("reporting_type") == "hebdomadaire":
    st.markdown('<div class="section-title">📅 Reporting Hebdomadaire</div>', unsafe_allow_html=True)

    uploaded_file = st.file_uploader("📁 Importer le fichier Excel brut (hebdomadaire)", type=["xlsx", "csv"])

    if uploaded_file:
        st.markdown("<br>", unsafe_allow_html=True)

        if uploaded_file.name.endswith('.csv'):
            df = pd.read_csv(uploaded_file, encoding='utf-8', sep=';')
        else:
            xls = pd.ExcelFile(uploaded_file)
            sheet_names = xls.sheet_names
            selected_sheet = st.selectbox("🗂 Choisir la feuille à exploiter :", options=sheet_names)
            df = pd.read_excel(uploaded_file, sheet_name=selected_sheet)

        st.markdown("<br>", unsafe_allow_html=True)

        vto_df = load_vto()
        logins_concernes = vto_df["LOGIN"].astype(str).str.lower().tolist()
        details = ["En Cours-Identification", "Identifie", "Identifie Photo"]

        column_mapping = {
            'MSISDN': 'TOTAL_SIM',
            'ACCUEIL_VENDEUR': 'PVT',
            'LOGIN_VENDEUR': 'LOGIN',
            'AGENCE_VENDEUR': 'DR'
        }

        missing_columns = [col for col in column_mapping.keys() if col not in df.columns]

        if missing_columns:
            st.error(f"❌ Colonnes manquantes : {', '.join(missing_columns)}")
            st.stop()

        df = df.rename(columns=column_mapping)

        if 'LOGIN' not in df.columns:
            st.error("❌ La colonne LOGIN n'a pas pu être créée.")
            st.stop()

        df['LOGIN'] = df['LOGIN'].astype(str).str.lower()
        df['DR'] = df['DR'].astype(str).str.strip().str.upper()
        df['NOM_VENDEUR'] = df['NOM_VENDEUR'].astype(str).str.strip().str.upper()
        df['PRENOM_VENDEUR'] = df['PRENOM_VENDEUR'].astype(str).str.strip().str.upper()

        df_filtre = df[df['LOGIN'].isin(logins_concernes) & df['ETAT_IDENTIFICATION'].astype(str).isin(details)]
        df_filtre["DR"] = df_filtre["DR"].replace(DRV_MAPPING)

        st.success(f"✅ Fichier filtré avec succès ! {df_filtre.shape[0]} ventes hebdomadaires")

        # Métriques
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-value">{df_filtre.shape[0]}</div>
                <div class="metric-label">📊 Ventes de la Semaine</div>
            </div>
            """, unsafe_allow_html=True)
        with col2:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-value">{df_filtre['LOGIN'].nunique()}</div>
                <div class="metric-label">👤 VTO Actifs</div>
            </div>
            """, unsafe_allow_html=True)
        with col3:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-value">{df_filtre['PVT'].nunique()}</div>
                <div class="metric-label">🏪 PVT Concernés</div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # 📊 TABLEAU 1 : Résumé par PVT
        st.markdown('<div class="section-title">📊 Résumé par PVT</div>', unsafe_allow_html=True)

        df_pvt_summary = df_filtre.groupby(['DR', 'PVT'], as_index=False).size()
        df_pvt_summary.columns = ['DR', 'PVT', 'TOTAL_SIM']
        df_pvt_summary['OBJECTIF'] = 240
        df_pvt_summary['TR'] = (df_pvt_summary['TOTAL_SIM'] / df_pvt_summary['OBJECTIF'] * 100).round(0).astype(int).astype(str) + '%'

        df_pvt_summary = df_pvt_summary.sort_values(['DR', 'PVT'])

        total_sim = df_pvt_summary['TOTAL_SIM'].sum()
        total_objectif = df_pvt_summary['OBJECTIF'].sum()
        total_tr = round((total_sim / total_objectif * 100), 1)

        df_pvt_summary_display = df_pvt_summary.copy()
        df_pvt_summary_display.loc[len(df_pvt_summary_display)] = ['', 'TOTAL', total_sim, total_objectif, f'{total_tr}%']

        st.dataframe(df_pvt_summary_display, use_container_width=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # 📊 TABLEAU 2 : Détails par VTO
        st.markdown('<div class="section-title">📋 Détails par VTO</div>', unsafe_allow_html=True)

        df_reporting = df_filtre.groupby(['DR', 'PVT', 'PRENOM_VENDEUR', 'NOM_VENDEUR', 'LOGIN'], as_index=False).size()
        df_reporting.columns = ['DR', 'PVT', 'PRENOM_VENDEUR', 'NOM_VENDEUR', 'LOGIN', 'TOTAL_SIM']
        df_reporting = df_reporting.sort_values(['DR', 'PVT', 'TOTAL_SIM'], ascending=[True, True, False])

        total_sim_vto = df_reporting['TOTAL_SIM'].sum()
        df_reporting_display = df_reporting.copy()
        df_reporting_display.loc[len(df_reporting_display)] = ['', '', '', '', 'TOTAL', total_sim_vto]

        st.dataframe(df_reporting_display, use_container_width=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # 🆕 SECTION SAUVEGARDE DANS LA BASE DE DONNÉES
        st.markdown('<div class="section-title">💾 Sauvegarde dans la Base de Données</div>', unsafe_allow_html=True)

        col_dates = st.columns(4)

        with col_dates[0]:
            date_debut = st.date_input(
                "📅 Date de début",
                value=datetime.now().date(),
                format="DD/MM/YYYY",
                key="date_debut_hebdo"
            )

        with col_dates[1]:
            date_fin = st.date_input(
                "📅 Date de fin",
                value=datetime.now().date(),
                format="DD/MM/YYYY",
                key="date_fin_hebdo"
            )

        with col_dates[2]:
            # Calculer le numéro de semaine
            semaine_num = date_debut.isocalendar()[1]
            annee = date_debut.year
            semaine_str = f"S{semaine_num:02d}-{annee}"
            st.text_input("Semaine", value=semaine_str, disabled=True)

        with col_dates[3]:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("💾 Sauvegarder", use_container_width=True, key="save_hebdo"):
                date_debut_str = date_debut.strftime('%Y-%m-%d')
                date_fin_str = date_fin.strftime('%Y-%m-%d')

                df_pvt_save = df_pvt_summary.copy()
                df_vto_save = df_reporting.copy()

                success, message = db.save_weekly_report(
                    semaine_str,
                    date_debut_str,
                    date_fin_str,
                    df_pvt_save,
                    df_vto_save
                )
                if success:
                    st.success(f"✅ {message}")
                else:
                    st.error(f"❌ {message}")

        # Option pour charger depuis la BDD
        if st.button("📂 Charger un reporting existant", use_container_width=True, key="show_load_hebdo"):
            st.session_state.show_load_hebdo = True

        if st.session_state.get('show_load_hebdo'):
            st.markdown("---")
            available_weeks = db.get_available_dates('weekly')

            if available_weeks:
                col_select, col_action, col_cancel = st.columns([3, 1, 1])

                with col_select:
                    week_options = {f"{w[0]} ({w[1]} au {w[2]})": w[0] for w in available_weeks}
                    selected_week_display = st.selectbox(
                        "Choisir une semaine",
                        options=list(week_options.keys()),
                        key="select_week_hebdo"
                    )
                    selected_week = week_options[selected_week_display]

                with col_action:
                    st.markdown("<br>", unsafe_allow_html=True)
                    if st.button("📥 Charger", key="confirm_load_hebdo"):
                        st.info(f"📊 Données de la semaine {selected_week} chargées depuis la BDD")
                        st.session_state.show_load_hebdo = False
                        st.rerun()

                with col_cancel:
                    st.markdown("<br>", unsafe_allow_html=True)
                    if st.button("❌ Annuler", key="cancel_load_hebdo"):
                        st.session_state.show_load_hebdo = False
                        st.rerun()
            else:
                st.info("Aucun reporting sauvegardé dans la base de données")
                if st.button("Fermer", key="close_load_hebdo"):
                    st.session_state.show_load_hebdo = False
                    st.rerun()

        st.markdown("<br>", unsafe_allow_html=True)

        # 📊 GRAPHIQUES - MÊME ORGANISATION QUE POUR LE JOURNALIER
        st.markdown('<div class="section-title">📊 VISUALISATION DES PERFORMANCES</div>', unsafe_allow_html=True)

        # Première ligne : Top DR (circulaire + barres VERTICALES)
        st.markdown('<h4 style="text-align: center; color: #FF7900;">🗺 Distribution par Direction Régionale</h4>', unsafe_allow_html=True)

        df_drv = df_reporting.groupby('DR').agg({'TOTAL_SIM': 'sum'}).reset_index()
        df_drv = df_drv.sort_values('TOTAL_SIM', ascending=False)

        total_ventes = df_drv['TOTAL_SIM'].sum()
        df_drv['POURCENTAGE'] = (df_drv['TOTAL_SIM'] / total_ventes * 100).round(1)

        col_dr1, col_dr2 = st.columns(2)

        with col_dr1:
            # Diagramme circulaire
            fig_pie_dr = px.pie(
                df_drv,
                values='TOTAL_SIM',
                names='DR',
                title='Distribution des ventes par DR (Hebdomadaire)',
                color_discrete_sequence=['#FF7900', '#FF5000', '#FF3000', '#E57200', '#CC6600', '#B35900'],
                hole=0.3
            )

            df_drv['LABEL'] = df_drv.apply(lambda row: f"{row['DR']}<br>{row['TOTAL_SIM']} ventes<br>({row['POURCENTAGE']}%)", axis=1)

            fig_pie_dr.update_traces(
                textposition='inside',
                textinfo='text',
                text=df_drv['LABEL'],
                textfont_size=10,
                marker=dict(line=dict(color='white', width=2)),
                hovertemplate="<b>%{label}</b><br>" +
                             "Ventes: %{value}<br>" +
                             "Pourcentage: %{percent}<br>" +
                             "<extra></extra>"
            )

            fig_pie_dr.update_layout(
                title=dict(
                    text="Distribution des ventes par DR (Hebdomadaire)",
                    font=dict(size=16, family='Poppins', color='#FF7900'),
                    x=0.5,
                    y=0.95
                ),
                font=dict(family='Poppins', size=10),
                height=400,
                margin=dict(t=70, b=30, l=30, r=30),
                showlegend=False,
                hoverlabel=dict(
                    bgcolor="white",
                    font_size=11,
                    font_family="Poppins"
                )
            )

            st.plotly_chart(fig_pie_dr, use_container_width=True)

            st.download_button(
                label="📥 Télécharger (PNG)",
                data=download_plotly_as_png(fig_pie_dr, "Distribution_DR_Circulaire_Hebdo"),
                file_name="Distribution_DR_Circulaire_Hebdo.png",
                mime="image/png",
                key="download_dr_pie_hebdo"
            )

        with col_dr2:
            # Diagramme à barres VERTICALES triées
            fig_bar_dr = go.Figure()

            df_drv_bar = df_drv.sort_values('TOTAL_SIM', ascending=False)

            fig_bar_dr.add_trace(go.Bar(
                x=df_drv_bar['DR'],
                y=df_drv_bar['TOTAL_SIM'],
                marker_color='#FF7900',
                text=df_drv_bar.apply(lambda row: f"{row['TOTAL_SIM']}", axis=1),
                textposition='outside',
                textfont=dict(size=10, color='#333'),
                textangle=0,
                hovertemplate="<b>%{x}</b><br>" +
                             "Ventes: %{y}<br>" +
                             "Pourcentage: %{customdata}%<br>" +
                             "<extra></extra>",
                customdata=df_drv_bar['POURCENTAGE']
            ))

            fig_bar_dr.update_layout(
                title=dict(
                    text="Performances par DR (Hebdomadaire)",
                    font=dict(size=16, family='Poppins', color='#FF7900'),
                    x=0.5,
                    y=0.95
                ),
                xaxis_title='',
                yaxis_title='Nombre de ventes',
                template='plotly_white',
                height=400,
                margin=dict(t=70, b=30, l=60, r=30),
                font=dict(family='Poppins', size=10),
                hoverlabel=dict(
                    bgcolor="white",
                    font_size=11,
                    font_family="Poppins"
                ),
                xaxis=dict(
                    tickangle=0,
                    tickfont=dict(size=10),
                    automargin=True,
                    showgrid=False,
                    categoryorder='total descending'
                ),
                yaxis=dict(
                    showgrid=True,
                    gridcolor='lightgray'
                )
            )

            st.plotly_chart(fig_bar_dr, use_container_width=True)

            st.download_button(
                label="📥 Télécharger (PNG)",
                data=download_plotly_as_png(fig_bar_dr, "Distribution_DR_Barres_Hebdo"),
                file_name="Distribution_DR_Barres_Hebdo.png",
                mime="image/png",
                key="download_dr_bar_hebdo"
            )

        # Deuxième ligne : Top PVT (Top 5) et Top VTO (Top 10)
        col_pvt, col_vto = st.columns(2)

        with col_pvt:
            st.markdown('<h4 style="text-align: center; color: #009CA6;">🏪 Top 5 Points de Vente</h4>', unsafe_allow_html=True)

            df_pvt_summary_chart = df_filtre.groupby(['DR', 'PVT']).agg({'TOTAL_SIM': 'sum'}).reset_index()
            df_top_pvt = df_pvt_summary_chart.nlargest(5, 'TOTAL_SIM')  # Top 5 seulement
            df_top_pvt = df_top_pvt.sort_values('TOTAL_SIM', ascending=True)

            def create_pvt_label_hebdo(pvt_name, dr, max_length=25):
                short_name = pvt_name[:max_length-3] + "..." if len(pvt_name) > max_length else pvt_name
                return f"{short_name} (DR: {dr})"

            df_top_pvt['LABEL'] = df_top_pvt.apply(
                lambda row: create_pvt_label_hebdo(row['PVT'], row['DR']),
                axis=1
            )

            fig_pvt = go.Figure()

            fig_pvt.add_trace(go.Bar(
                y=df_top_pvt['LABEL'],
                x=df_top_pvt['TOTAL_SIM'],
                orientation='h',
                marker_color='#009CA6',
                # SUPPRIMÉ: text=df_top_pvt['TOTAL_SIM'],
                # SUPPRIMÉ: textposition='outside',
                # SUPPRIMÉ: textfont=dict(size=10, color='#333'),
                # SUPPRIMÉ: textangle=0,
                hovertemplate="<b>%{customdata[0]}</b><br>" +
                             "Ventes: %{x}<br>" +
                             "DR: %{customdata[1]}<br>" +
                             "<extra></extra>",
                customdata=df_top_pvt[['PVT', 'DR']]
            ))

            fig_pvt.update_layout(
                title=dict(
                    text="Top 5 PVT - Performances Hebdomadaires",
                    font=dict(size=16, family='Poppins', color='#009CA6'),
                    x=0.5,
                    y=0.95
                ),
                xaxis_title='Nombre de ventes',
                yaxis_title='',
                template='plotly_white',
                height=350,  # Hauteur réduite pour 5 éléments
                margin=dict(t=70, b=30, l=250, r=30),
                font=dict(family='Poppins', size=10),
                hoverlabel=dict(
                    bgcolor="white",
                    font_size=11,
                    font_family="Poppins"
                ),
                yaxis=dict(
                    tickfont=dict(size=9),
                    automargin=True,
                    showgrid=False
                ),
                xaxis=dict(
                    showgrid=True,
                    gridcolor='lightgray'
                )
            )

            st.plotly_chart(fig_pvt, use_container_width=True)

            st.download_button(
                label="📥 Télécharger (PNG)",
                data=download_plotly_as_png(fig_pvt, "Top_5_PVT_Hebdo"),
                file_name="Top_5_PVT_Hebdo.png",
                mime="image/png",
                key="download_pvt_chart_hebdo"
            )

        with col_vto:
            st.markdown('<h4 style="text-align: center; color: #FF5000;">👥 Top 10 Vendeurs (VTO)</h4>', unsafe_allow_html=True)

            df_top10 = df_reporting.nlargest(10, 'TOTAL_SIM').copy()
            df_top10 = df_top10.sort_values('TOTAL_SIM', ascending=False)

            def create_vto_label_hebdo(prenom, nom, pvt, dr, max_length=20):
                nom_complet = f"{prenom} {nom}"
                if len(nom_complet) > max_length:
                    nom_complet = nom_complet[:max_length-3] + "..."

                return f"{nom_complet}<br>(PVT: {pvt[:12]}...)" if len(pvt) > 12 else f"{nom_complet}<br>(PVT: {pvt})"

            df_top10['LABEL'] = df_top10.apply(
                lambda row: create_vto_label_hebdo(
                    row['PRENOM_VENDEUR'],
                    row['NOM_VENDEUR'],
                    row['PVT'],
                    row['DR']
                ),
                axis=1
            )

            fig_top10 = go.Figure()

            fig_top10.add_trace(go.Bar(
                x=df_top10['LABEL'],
                y=df_top10['TOTAL_SIM'],
                marker_color='#FF5000',
                text=df_top10['TOTAL_SIM'],
                textposition='outside',
                textfont=dict(size=10, color='#333'),
                marker_line=dict(color='white', width=1),
                hovertemplate="<b>%{customdata[0]} %{customdata[1]}</b><br>" +
                             "Ventes: %{y}<br>" +
                             "PVT: %{customdata[2]}<br>" +
                             "DR: %{customdata[3]}<br>" +
                             "<extra></extra>",
                customdata=df_top10[['PRENOM_VENDEUR', 'NOM_VENDEUR', 'PVT', 'DR']]
            ))

            fig_top10.update_layout(
                title=dict(
                    text="Top 10 Vendeurs (VTO) - Hebdomadaire",
                    font=dict(size=16, family='Poppins', color='#FF5000'),
                    x=0.5,
                    y=0.95
                ),
                xaxis_title='',
                yaxis_title='Nombre de ventes',
                template='plotly_white',
                height=400,
                margin=dict(t=70, b=100, l=60, r=30),
                font=dict(family='Poppins', size=10),
                hoverlabel=dict(
                    bgcolor="white",
                    font_size=11,
                    font_family="Poppins"
                ),
                xaxis=dict(
                    tickangle=-35,
                    tickfont=dict(size=9),
                    automargin=True,
                    showgrid=False
                ),
                yaxis=dict(
                    showgrid=True,
                    gridcolor='lightgray'
                )
            )

            st.plotly_chart(fig_top10, use_container_width=True)

            st.download_button(
                label="📥 Télécharger ce graphique (PNG pour PowerPoint)",
                data=download_plotly_as_png(fig_top10, "Top_VTO_Hebdo"),
                file_name="Top_VTO_Hebdo.png",
                mime="image/png",
                key="download_top10_hebdo"
            )

        # 🔹 MÉTRIQUES DE PERFORMANCE
        st.markdown('<br><br>', unsafe_allow_html=True)
        st.markdown('<div class="section-title">📈 RÉSUMÉ DES PERFORMANCES</div>', unsafe_allow_html=True)

        col1, col2, col3, col4 = st.columns(4)

        with col1:
            st.markdown(f"""
            <div style="
                background: white;
                border-radius: 12px;
                padding: 1.2rem;
                box-shadow: 0 4px 15px rgba(255, 121, 0, 0.1);
                border-left: 5px solid #FF7900;
                text-align: center;
            ">
                <div style="font-size: 2rem; font-weight: 700; color: #FF7900;">{df_filtre.shape[0]}</div>
                <div style="font-size: 0.9rem; color: #666; margin-top: 0.5rem;">📊 Total Ventes</div>
            </div>
            """, unsafe_allow_html=True)

        with col2:
            meilleur_dr = df_drv.nlargest(1, 'TOTAL_SIM')
            meilleur_dr_nom = meilleur_dr['DR'].iloc[0] if not meilleur_dr.empty else "N/A"
            meilleur_dr_ventes = meilleur_dr['TOTAL_SIM'].iloc[0] if not meilleur_dr.empty else 0

            st.markdown(f"""
            <div style="
                background: white;
                border-radius: 12px;
                padding: 1.2rem;
                box-shadow: 0 4px 15px rgba(0, 156, 166, 0.1);
                border-left: 5px solid #009CA6;
                text-align: center;
            ">
                <div style="font-size: 1.8rem; font-weight: 700; color: #009CA6;">{meilleur_dr_nom}</div>
                <div style="font-size: 0.9rem; color: #666; margin-top: 0.5rem;">🏆 Meilleure DR</div>
            </div>
            """, unsafe_allow_html=True)

        with col3:
            meilleur_pvt = df_top_pvt.nlargest(1, 'TOTAL_SIM')
            meilleur_pvt_nom = meilleur_pvt['PVT'].iloc[0][:15] + "..." if not meilleur_pvt.empty and len(meilleur_pvt['PVT'].iloc[0]) > 15 else (meilleur_pvt['PVT'].iloc[0] if not meilleur_pvt.empty else "N/A")
            meilleur_pvt_ventes = meilleur_pvt['TOTAL_SIM'].iloc[0] if not meilleur_pvt.empty else 0

            st.markdown(f"""
            <div style="
                background: white;
                border-radius: 12px;
                padding: 1.2rem;
                box-shadow: 0 4px 15px rgba(0, 212, 170, 0.1);
                border-left: 5px solid #00D4AA;
                text-align: center;
            ">
                <div style="font-size: 1.8rem; font-weight: 700; color: #00D4AA;">{meilleur_pvt_nom}</div>
                <div style="font-size: 0.9rem; color: #666; margin-top: 0.5rem;">🏪 Meilleur PVT</div>
            </div>
            """, unsafe_allow_html=True)

        with col4:
            meilleur_vto = df_top10.nlargest(1, 'TOTAL_SIM')
            meilleur_vto_nom = meilleur_vto['PRENOM_VENDEUR'].iloc[0][:8] + "..." if not meilleur_vto.empty and len(meilleur_vto['PRENOM_VENDEUR'].iloc[0]) > 8 else (meilleur_vto['PRENOM_VENDEUR'].iloc[0] if not meilleur_vto.empty else "N/A")
            meilleur_vto_ventes = meilleur_vto['TOTAL_SIM'].iloc[0] if not meilleur_vto.empty else 0

            st.markdown(f"""
            <div style="
                background: white;
                border-radius: 12px;
                padding: 1.2rem;
                box-shadow: 0 4px 15px rgba(255, 80, 0, 0.1);
                border-left: 5px solid #FF5000;
                text-align: center;
            ">
                <div style="font-size: 1.8rem; font-weight: 700; color: #FF5000;">{meilleur_vto_nom}</div>
                <div style="font-size: 0.9rem; color: #666; margin-top: 0.5rem;">👤 Meilleur VTO</div>
            </div>
            """, unsafe_allow_html=True)