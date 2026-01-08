import sys
import os
import subprocess

# Installer kaleido si nécessaire
try:
    import kaleido
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "kaleido", "--break-system-packages"])
    import kaleido

# CORRECTION ICI : __file__ au lieu de _file_
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import plotly.express as px
import plotly.graph_objects as go
from PIL import Image
from pathlib import Path

from utils import load_vto, load_pvt

# ====================
# Configuration page
# ====================
st.set_page_config(page_title="LOUMA - Paiement Mensuel", layout="wide", initial_sidebar_state="expanded")

# ====================
# CSS personnalisé Orange Sonatel
# ====================
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600;700&display=swap');

    header[data-testid="stHeader"] { display: none; }

    .block-container {
        padding-top: 2rem !important;
        padding-bottom: 2rem !important;
    }

    .main {
        font-family: 'Poppins', sans-serif;
        background: linear-gradient(135deg, #fff5f0 0%, #ffffff 50%, #f0f8ff 100%);
    }

    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #FF7900 0%, #FF5000 100%) !important;
    }

    section[data-testid="stSidebar"] * {
        color: white !important;
    }

    .stDataFrame {
        border-radius: 15px;
        overflow: hidden;
        box-shadow: 0 8px 20px rgba(255, 121, 0, 0.15);
        border: 2px solid #FFE5CC;
    }

    .section-title {
        background: linear-gradient(135deg, #FF7900 0%, #FF5000 100%);
        color: white;
        padding: 0.7rem 1.2rem;
        border-radius: 10px;
        font-weight: 600;
        font-size: 1.1rem;
        margin-bottom: 0.8rem;
        box-shadow: 0 4px 12px rgba(255, 121, 0, 0.25);
        text-align: center;
        max-width: 500px;
        margin-left: auto;
        margin-right: auto;
    }

    .stButton > button {
        background: linear-gradient(135deg, #FF7900 0%, #FF5000 100%);
        color: white;
        border: none;
        border-radius: 10px;
        padding: 0.8rem 2rem;
        font-weight: 600;
        font-size: 1.1rem;
        box-shadow: 0 4px 12px rgba(255, 121, 0, 0.3);
        transition: all 0.3s ease;
        width: 100%;
    }

    .stButton > button:hover {
        background: linear-gradient(135deg, #FF5000 0%, #FF3000 100%);
        box-shadow: 0 6px 18px rgba(255, 121, 0, 0.5);
        transform: translateY(-2px);
    }

    .stDownloadButton > button {
        background: linear-gradient(135deg, #00D4AA 0%, #00B890 100%);
        color: white;
        border: none;
        border-radius: 10px;
        padding: 0.8rem 2rem;
        font-weight: 600;
        font-size: 1.1rem;
        box-shadow: 0 4px 12px rgba(0, 212, 170, 0.3);
        transition: all 0.3s ease;
        width: 100%;
    }

    .stDownloadButton > button:hover {
        background: linear-gradient(135deg, #00B890 0%, #009A7A 100%);
        box-shadow: 0 6px 18px rgba(0, 212, 170, 0.5);
        transform: translateY(-2px);
    }

    .metric-card {
        background: white;
        border-radius: 12px;
        padding: 1.5rem;
        box-shadow: 0 4px 15px rgba(255, 121, 0, 0.15);
        border: 2px solid #FFE5CC;
        text-align: center;
    }

    .metric-value {
        font-size: 2rem;
        font-weight: 700;
        color: #FF7900;
    }

    .metric-label {
        font-size: 1rem;
        color: #666;
        margin-top: 0.5rem;
    }

    .info-card {
        background: linear-gradient(135deg, #009CA6 0%, #00B8C5 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 12px;
        box-shadow: 0 4px 15px rgba(0, 156, 166, 0.3);
        margin-bottom: 1.5rem;
    }

    @keyframes fadeIn {
        from { opacity: 0; transform: translateY(20px); }
        to { opacity: 1; transform: translateY(0); }
    }

    .stDataFrame, .metric-card, .info-card {
        animation: fadeIn 0.5s ease-out;
    }
</style>
""", unsafe_allow_html=True)

# ====================
# Charger le logo
# ====================
logo = None
parent_dir = Path(__file__).parent.parent
logo_paths = [
    parent_dir / "assets" / "logo sonatel.png",
    Path("assets") / "logo sonatel.png",
    Path("../assets/logo sonatel.png"),
]

for logo_path in logo_paths:
    try:
        if logo_path.exists():
            logo = Image.open(logo_path)
            break
    except:
        continue

# ====================
# Header avec logo et titre
# ====================
col_logo, col_title = st.columns([1, 3])

with col_logo:
    if logo:
        st.image(logo, width=280)

with col_title:
    st.markdown("""
    <div style="
        background: linear-gradient(135deg, #FF7900 0%, #FF5000 100%);
        padding: 2rem;
        border-radius: 20px;
        box-shadow: 0 8px 25px rgba(255, 121, 0, 0.4);
        display: flex;
        flex-direction: column;
        justify-content: center;
        border: 3px solid rgba(255, 255, 255, 0.2);
        height: 100%;
    ">
        <h1 style="
            color: white;
            font-size: 2.5rem;
            font-weight: 700;
            margin: 0;
            text-shadow: 3px 3px 10px rgba(0, 0, 0, 0.3);
        ">
            💰 Paiement Mensuel Global
        </h1>
        <p style="
            color: rgba(255, 255, 255, 0.95);
            font-size: 1.2rem;
            margin: 0.8rem 0 0 0;
            font-weight: 400;
        ">
            SIM + OM - Orange Sénégal
        </p>
    </div>
    """, unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# ====================
# Upload des fichiers
# ====================
st.markdown('<div class="section-title">📁 Import des Fichiers</div>', unsafe_allow_html=True)

col1, col2 = st.columns(2)
with col1:
    file_sim = st.file_uploader("📥 Importer le fichier SIM", type=["xlsx", "csv"])
with col2:
    file_om = st.file_uploader("📥 Importer le fichier OM", type=["xlsx", "csv"])

if file_sim and file_om:
    st.markdown("<br>", unsafe_allow_html=True)

    # === TRAITEMENT FICHIER SIM ===
    if file_sim.name.endswith(".csv"):
        df_sim = pd.read_csv(file_sim, sep=";", encoding="utf-8")
    else:
        xls = pd.ExcelFile(file_sim)
        sheet_names = xls.sheet_names
        selected_sheet = st.selectbox("🗂 Feuille SIM :", sheet_names, key="sim_sheet")
        df_sim = pd.read_excel(file_sim, sheet_name=selected_sheet)

    # === TRAITEMENT FICHIER OM ===
    if file_om.name.endswith(".csv"):
        df_om = pd.read_csv(file_om, sep=";", encoding="utf-8")
    else:
        xls = pd.ExcelFile(file_om)
        sheet_names = xls.sheet_names
        selected_sheet = st.selectbox("🗂 Feuille OM :", sheet_names, key="om_sheet")
        df_om = pd.read_excel(file_om, sheet_name=selected_sheet)

    # === CHARGEMENT VTO ===
    vto_df = load_vto()
    vto_df['LOGIN'] = vto_df['LOGIN'].astype(str).str.strip().str.lower()
    logins_concernes = vto_df["LOGIN"].astype(str).tolist()
    details = ["En Cours-Identification", "Identifie", "Identifie Photo"]

    # === TRAITEMENT SIM ===
    df_sim['LOGIN_VENDEUR'] = df_sim['LOGIN_VENDEUR'].astype(str).str.strip().str.lower()
    df = df_sim.copy()

    df = df.rename(columns={
        'MSISDN': 'REALISATION_SIM',
        'ACCUEIL_VENDEUR': 'PVT',
        'LOGIN_VENDEUR': 'LOGIN',
        'AGENCE_VENDEUR': 'DRV'
    })

    df['LOGIN'] = df['LOGIN'].astype(str)
    df['DRV'] = df['DRV'].astype(str).str.strip().str.upper()
    df['NOM_VENDEUR'] = df['NOM_VENDEUR'].astype(str).str.strip().str.upper()
    df['PRENOM_VENDEUR'] = df['PRENOM_VENDEUR'].astype(str).str.strip().str.upper()

    # Filtrer avant le groupby
    df_filtre = df[df['LOGIN'].isin(logins_concernes) & df['ETAT_IDENTIFICATION'].astype(str).isin(details)]

    # Vérifier les duplications AVANT le groupby
    st.write("🔍 **Diagnostic des données SIM avant traitement:**")
    duplicates_check = df_filtre.groupby(['LOGIN', 'PVT']).size().reset_index(name='count')
    duplicates_found = duplicates_check[duplicates_check['count'] > 1]

    if not duplicates_found.empty:
        st.warning(f"⚠️ {len(duplicates_found)} combinaisons LOGIN+PVT apparaissent plusieurs fois dans les données sources")
        with st.expander("Voir les doublons détectés"):
            st.dataframe(duplicates_found.sort_values('count', ascending=False))

    df_filtre["DRV"] = df_filtre["DRV"].replace({
        "DV-DRV2_DIRECTION REGIONALE DES VENTES DAKAR 2": "DR2",
        "DV-DRVS_DIRECTION REGIONALE DES VENTES SUD": "DR SUD",
        "DV-DRVSE_DIRECTION REGIONALE DES VENTES SUD-EST": "SUD EST",
        "DV-DRVN_DIRECTION REGIONALE DES VENTES NORD": "DR NORD",
        "DV-DRVC_DIRECTION REGIONALE DES VENTES CENTRE": "DR CENTRE",
        "DV-DRVE_DIRECTION REGIONALE DES VENTES EST": "DR EST"
    })

    # GROUPBY avec agrégation - Cela devrait éliminer les duplications
    df_filtre = df_filtre.groupby(['DRV', 'PVT', 'PRENOM_VENDEUR', 'NOM_VENDEUR', 'LOGIN']).agg({
        'REALISATION_SIM': 'count'
    }).reset_index().sort_values(['DRV', 'PVT'])

    # Vérification APRÈS le groupby
    duplicates_after = df_filtre.groupby(['LOGIN', 'PVT']).size().reset_index(name='count')
    duplicates_after_found = duplicates_after[duplicates_after['count'] > 1]

    if not duplicates_after_found.empty:
        st.error(f"❌ PROBLÈME: {len(duplicates_after_found)} doublons persistent après le groupby!")
        with st.expander("Voir les doublons persistants"):
            st.dataframe(duplicates_after_found)
            # Afficher les lignes concernées
            problematic_logins = duplicates_after_found['LOGIN'].tolist()
            st.write("**Lignes problématiques:**")
            st.dataframe(df_filtre[df_filtre['LOGIN'].isin(problematic_logins)].sort_values(['LOGIN', 'PVT']))

    df_filtre['OBJECTIF SIM'] = 240
    df_filtre["TAUX D'ATTEINTE SIM"] = (df_filtre['REALISATION_SIM'] / df_filtre['OBJECTIF SIM']).apply(lambda x: f"{round(x*100)}%")
    df_filtre['SI 100% ATTEINT SIM'] = 75000
    df_filtre['PAIEMENT_SIM'] = df_filtre['REALISATION_SIM'].apply(lambda x: 75000 if x >= 240 else round((x/240)*75000))

    df_filtre = df_filtre.merge(vto_df[["LOGIN", "KABBU"]], how="left")

    # === TRAITEMENT OM ===
    df_om['LOGIN'] = df_om['LOGIN'].astype(str).str.strip().str.lower()
    df_om['NOM_VENDEUR'] = df_om['NOM_VENDEUR'].astype(str).str.strip().str.upper()
    df_om['PRENOM_VENDEUR'] = df_om['PRENOM_VENDEUR'].astype(str).str.strip().str.upper()

    df_filtre_om = df_om[df_om['LOGIN'].isin(logins_concernes)]
    df_filtre_om = df_filtre_om.fillna(0)

    # Vérification des doublons OM
    duplicates_om_check = df_filtre_om.groupby('LOGIN').size().reset_index(name='count')
    duplicates_om_found = duplicates_om_check[duplicates_om_check['count'] > 1]

    if not duplicates_om_found.empty:
        st.warning(f"⚠️ {len(duplicates_om_found)} LOGIN apparaissent plusieurs fois dans les données OM")

    df_filtre_om['OBJECTIF OM'] = 120
    df_filtre_om["TAUX D'ATTEINTE OM"] = ((df_filtre_om['REALISATION_OM'] / df_filtre_om['OBJECTIF OM']).fillna(0).apply(lambda x: f"{round(x*100)}%"))
    df_filtre_om['SI 100% ATTEINT OM'] = 25000
    df_filtre_om['PAIEMENT_OM'] = df_filtre_om['REALISATION_OM'].apply(lambda x: 25000 if x >= 120 else round((x/120)*25000))

    df_filtre_om = df_filtre_om.merge(vto_df[["LOGIN", "DRV", "PVT"]], how="left")

    # === FUSION SIM + OM ===
    st.write("🔄 **Fusion des données SIM et OM...**")

    # Déduplication STRICTE avant la fusion
    nb_lignes_sim_avant = len(df_filtre)
    df_filtre_unique = df_filtre.drop_duplicates(subset=['LOGIN', 'PVT'], keep='first')
    nb_lignes_sim_apres = len(df_filtre_unique)

    if nb_lignes_sim_avant > nb_lignes_sim_apres:
        st.success(f"✅ SIM: {nb_lignes_sim_avant - nb_lignes_sim_apres} doublons supprimés ({nb_lignes_sim_avant} → {nb_lignes_sim_apres} lignes)")

    nb_lignes_om_avant = len(df_filtre_om)
    df_filtre_om_unique = df_filtre_om.drop_duplicates(subset=['LOGIN'], keep='first')
    nb_lignes_om_apres = len(df_filtre_om_unique)

    if nb_lignes_om_avant > nb_lignes_om_apres:
        st.success(f"✅ OM: {nb_lignes_om_avant - nb_lignes_om_apres} doublons supprimés ({nb_lignes_om_avant} → {nb_lignes_om_apres} lignes)")

    # MERGE avec données nettoyées
    df_test = pd.merge(
        df_filtre_unique,
        df_filtre_om_unique[["LOGIN", "REALISATION_OM", "OBJECTIF OM", "TAUX D'ATTEINTE OM", "SI 100% ATTEINT OM", "PAIEMENT_OM"]],
        on=["LOGIN"],
        how="outer",
        suffixes=('', '_om')
    )

    # Vérification finale après le merge
    final_duplicates = df_test.groupby(['LOGIN', 'PVT']).size().reset_index(name='count')
    final_duplicates_found = final_duplicates[final_duplicates['count'] > 1]

    if not final_duplicates_found.empty:
        st.error(f"❌ {len(final_duplicates_found)} doublons dans le résultat final!")
        with st.expander("Voir les doublons finaux"):
            st.dataframe(final_duplicates_found)
    else:
        st.success("✅ Aucun doublon dans le résultat final!")

    df_test["PAIEMENT CHAUFFEUR"] = None
    df_test["PAIEMENT SIM + OM + CHAUFFEUR"] = None

    # Remplir les valeurs manquantes
    df_test['PAIEMENT_SIM'] = df_test['PAIEMENT_SIM'].fillna(0)
    df_test['PAIEMENT_OM'] = df_test['PAIEMENT_OM'].fillna(0)
    df_test['REALISATION_SIM'] = df_test['REALISATION_SIM'].fillna(0)
    df_test['REALISATION_OM'] = df_test['REALISATION_OM'].fillna(0)

    # === CRÉATION DES TOTAUX PAR PVT ET DRV ===
    df_test_with_totals = pd.DataFrame(columns=df_test.columns)

    for drv, group_drv in df_test.groupby('DRV'):
        for pvt, group_pvt in group_drv.groupby('PVT'):
            df_test_with_totals = pd.concat([df_test_with_totals, group_pvt], ignore_index=True)

            total_paiement_om = group_pvt['PAIEMENT_OM'].sum()
            total_sim = group_pvt['REALISATION_SIM'].sum()
            total_obj = group_pvt['OBJECTIF SIM'].sum()
            si_total_atteint = group_pvt['SI 100% ATTEINT SIM'].sum()
            tr_mean = group_pvt["TAUX D'ATTEINTE SIM"].apply(lambda x: float(x.strip('%'))).mean()
            total_om = group_pvt['REALISATION_OM'].sum()
            total_obj_om = group_pvt['OBJECTIF OM'].sum()
            si_total_atteint_om = group_pvt['SI 100% ATTEINT OM'].sum()
            tr_mean_om = (group_pvt["TAUX D'ATTEINTE OM"].apply(lambda x: float(str(x).replace('%', '').strip()) if pd.notnull(x) else 0).mean())
            total_paiement_sim = group_pvt['PAIEMENT_SIM'].sum()
            chauffeur = 100000
            total_pvt = total_paiement_sim + chauffeur + total_paiement_om

            row_total = {
                'PVT': "TOTAL PVT",
                'REALISATION_SIM': total_sim,
                'OBJECTIF SIM': total_obj,
                "TAUX D'ATTEINTE SIM": f'{tr_mean:.1f}%',
                'SI 100% ATTEINT SIM': si_total_atteint,
                'REALISATION_OM': total_om,
                'OBJECTIF OM': total_obj_om,
                "TAUX D'ATTEINTE OM": f'{tr_mean_om:.1f}%',
                'SI 100% ATTEINT OM': si_total_atteint_om,
                'PAIEMENT_OM': total_paiement_om,
                'PAIEMENT_SIM': total_paiement_sim,
                'PAIEMENT CHAUFFEUR': chauffeur,
                'PAIEMENT SIM + OM + CHAUFFEUR': total_pvt
            }
            df_test_with_totals = pd.concat([df_test_with_totals, pd.DataFrame([row_total])], ignore_index=True)

        total_paiement_om_drv = group_drv['PAIEMENT_OM'].sum()
        total_paiement_sim_drv = group_drv['PAIEMENT_SIM'].sum()
        chauffeur_drv = 200000
        total = chauffeur_drv + total_paiement_om_drv + total_paiement_sim_drv

        row_total_drv = {
            'DRV': f"{drv}",
            'PVT': "TOTAL",
            'PAIEMENT_OM': total_paiement_om_drv,
            'PAIEMENT_SIM': total_paiement_sim_drv,
            'PAIEMENT CHAUFFEUR': chauffeur_drv,
            'PAIEMENT SIM + OM + CHAUFFEUR': total
        }
        df_test_with_totals = pd.concat([df_test_with_totals, pd.DataFrame([row_total_drv])], ignore_index=True)

    df_test_with_totals = df_test_with_totals.rename(columns={
        'PRENOM_VENDEUR': 'PRENOM_VENDEUR',
        'NOM_VENDEUR': 'NOM_VENDEUR'
    })

    # === CALCUL MONTANTS ===
    df_test["MONTANT"] = df_test["PAIEMENT_SIM"] + df_test["PAIEMENT_OM"]

    df_par_pvt = df_test.groupby(["DRV", "PVT"]).agg({'MONTANT': 'sum'}).reset_index()
    df_par_pvt["MONTANT"] = df_par_pvt["MONTANT"] + 100000
    df_par_pvt["GAIN PVT (5%)"] = df_par_pvt["MONTANT"] * 0.05
    df_par_pvt["TOTAL GENERAL"] = df_par_pvt["MONTANT"] + df_par_pvt["GAIN PVT (5%)"]

    pvt_df = load_pvt()
    df_par_pvt = df_par_pvt.merge(pvt_df[["PVT", "CONTACT"]], on="PVT", how="left")
    df_par_pvt = df_par_pvt[["DRV", "PVT", "CONTACT", "MONTANT", "GAIN PVT (5%)", "TOTAL GENERAL"]]

    montant_sum = df_par_pvt['MONTANT'].sum()
    gain_sum = df_par_pvt['GAIN PVT (5%)'].sum()
    total_sum = df_par_pvt['TOTAL GENERAL'].sum()

    df_par_pvt_display = df_par_pvt.copy()
    df_par_pvt_display.loc[len(df_par_pvt_display)] = ['TOTAL', '', '', montant_sum, gain_sum, total_sum]

    # === AFFICHAGE DES MÉTRIQUES ===
    st.success("✅ Fichiers traités avec succès !")

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{df_test['LOGIN'].nunique()}</div>
            <div class="metric-label">👤 VTO Actifs</div>
        </div>
        """, unsafe_allow_html=True)
    with col2:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{df_test['PVT'].nunique()}</div>
            <div class="metric-label">🏪 PVT Concernés</div>
        </div>
        """, unsafe_allow_html=True)
    with col3:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{int(df_test['PAIEMENT_SIM'].sum()):,} FCFA</div>
            <div class="metric-label">💳 Total Paiement SIM</div>
        </div>
        """, unsafe_allow_html=True)
    with col4:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{int(df_test['PAIEMENT_OM'].sum()):,} FCFA</div>
            <div class="metric-label">💰 Total Paiement OM</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # === GRAPHIQUES ===
    st.markdown('<div class="section-title">📊 Analyse des Paiements</div>', unsafe_allow_html=True)

    # Graphique : Répartition SIM vs OM
    col1, col2 = st.columns(2)

    with col1:
        fig_pie = go.Figure(data=[go.Pie(
            labels=['Paiement SIM', 'Paiement OM'],
            values=[df_test['PAIEMENT_SIM'].sum(), df_test['PAIEMENT_OM'].sum()],
            hole=0.4,
            marker=dict(colors=['#FF7900', '#00D4AA'])
        )])

        fig_pie.update_layout(
            title='Répartition SIM vs OM',
            font=dict(family='Poppins', size=12),
            height=400
        )
        st.plotly_chart(fig_pie, use_container_width=True)

        try:
            buffer_pie = BytesIO()
            fig_pie.write_image(buffer_pie, format='png', width=800, height=600, scale=2)
            buffer_pie.seek(0)
            st.download_button(
                label="📥 Exporter Répartition SIM/OM",
                data=buffer_pie,
                file_name="repartition_sim_om.png",
                mime="image/png",
                key="export_pie"
            )
        except:
            st.warning("⚠ Kaleido requis pour l'export d'images")

    with col2:
        df_drv_paiement = df_test.groupby('DRV').agg({
            'PAIEMENT_SIM': 'sum',
            'PAIEMENT_OM': 'sum'
        }).reset_index()

        fig_bar = go.Figure()
        fig_bar.add_trace(go.Bar(
            x=df_drv_paiement['DRV'],
            y=df_drv_paiement['PAIEMENT_SIM'],
            name='SIM',
            marker_color='#FF7900'
        ))
        fig_bar.add_trace(go.Bar(
            x=df_drv_paiement['DRV'],
            y=df_drv_paiement['PAIEMENT_OM'],
            name='OM',
            marker_color='#00D4AA'
        ))

        fig_bar.update_layout(
            title='Paiements par DRV',
            xaxis_title='Direction Régionale',
            yaxis_title='Montant (FCFA)',
            barmode='group',
            template='plotly_white',
            font=dict(family='Poppins', size=12),
            height=400
        )
        st.plotly_chart(fig_bar, use_container_width=True)

        try:
            buffer_bar = BytesIO()
            fig_bar.write_image(buffer_bar, format='png', width=800, height=600, scale=2)
            buffer_bar.seek(0)
            st.download_button(
                label="📥 Exporter Paiements DRV",
                data=buffer_bar,
                file_name="paiements_drv.png",
                mime="image/png",
                key="export_bar"
            )
        except:
            pass

    st.markdown("<br>", unsafe_allow_html=True)

    # === TOP 10 VTO ===
    st.markdown('<div class="section-title">🏆 Top 10 VTO - Paiements Mensuels</div>', unsafe_allow_html=True)

    # Créer df_top10 à partir des données avant les totaux (df_test sans les lignes TOTAL)
    df_for_top10 = df_test[df_test['PVT'].notna() & (df_test['PVT'] != 'TOTAL PVT') & (df_test['PVT'] != 'TOTAL')].copy()
    df_for_top10['TOTAL_PAIEMENT'] = df_for_top10['PAIEMENT_SIM'] + df_for_top10['PAIEMENT_OM']
    df_top10 = df_for_top10.nlargest(10, 'TOTAL_PAIEMENT')

    # Vérifier si les colonnes existent avant de créer NOM_COMPLET
    if 'PRENOM_VENDEUR' in df_top10.columns and 'NOM_VENDEUR' in df_top10.columns:
        df_top10['NOM_COMPLET'] = df_top10['PRENOM_VENDEUR'].fillna('') + ' ' + df_top10['NOM_VENDEUR'].fillna('')
    else:
        df_top10['NOM_COMPLET'] = df_top10['LOGIN']

    fig_top10 = go.Figure()
    fig_top10.add_trace(go.Bar(
        x=df_top10['NOM_COMPLET'],
        y=df_top10['TOTAL_PAIEMENT'],
        marker_color='#FF5000',
        text=df_top10['TOTAL_PAIEMENT'].apply(lambda x: f"{int(x):,}"),
        textposition='outside'
    ))

    fig_top10.update_layout(
        title='Top 10 des VTO les Mieux Payés',
        xaxis_title='VTO',
        yaxis_title='Montant Total (FCFA)',
        template='plotly_white',
        height=500,
        font=dict(family='Poppins', size=12)
    )
    fig_top10.update_xaxes(tickangle=-45)
    st.plotly_chart(fig_top10, use_container_width=True)

    try:
        buffer_top10 = BytesIO()
        fig_top10.write_image(buffer_top10, format='png', width=1200, height=600, scale=2)
        buffer_top10.seek(0)
        st.download_button(
            label="📥 Exporter Top 10 VTO",
            data=buffer_top10,
            file_name="top10_vto_paiement.png",
            mime="image/png"
        )
    except:
        pass

    st.markdown("<br>", unsafe_allow_html=True)

    # === TABLEAUX ===
    st.markdown('<div class="section-title">📋 Résumé par PVT</div>', unsafe_allow_html=True)
    st.dataframe(df_par_pvt_display, use_container_width=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="section-title">📋 Détails Complets</div>', unsafe_allow_html=True)
    st.dataframe(df_test_with_totals, use_container_width=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # === EXPORT EXCEL FORMATÉ ===
    st.markdown('<div class="section-title">📥 Téléchargement du Rapport Excel</div>', unsafe_allow_html=True)

    try:
        buffer_output = BytesIO()

        with pd.ExcelWriter(buffer_output, engine='openpyxl') as writer:
            df_test_with_totals.to_excel(writer, sheet_name='Détails Paiement', index=False)
            df_par_pvt_display.to_excel(writer, sheet_name='Résumé PVT', index=False)

        buffer_output.seek(0)
        wb = load_workbook(buffer_output)

        # Style commun
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_font = Font(bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Formater "Détails Paiement"
        ws1 = wb['Détails Paiement']
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Fusionner DRV et PVT pour Détails Paiement
        drv_ranges = []
        current_drv = None
        drv_start = 2

        for row_idx in range(2, ws1.max_row + 1):
            drv_value = ws1.cell(row_idx, 1).value
            if drv_value and drv_value != current_drv:
                if current_drv is not None and row_idx > drv_start:
                    drv_ranges.append((drv_start, row_idx - 1, current_drv))
                current_drv = drv_value
                drv_start = row_idx

        if ws1.max_row >= drv_start:
            drv_ranges.append((drv_start, ws1.max_row, current_drv))

        for start_row, end_row, drv_value in drv_ranges:
            if end_row > start_row:
                ws1.merge_cells(f'A{start_row}:A{end_row}')
                ws1.cell(start_row, 1).alignment = Alignment(horizontal='left', vertical='center')
                ws1.cell(start_row, 1).font = Font(bold=True, size=10)

        # Appliquer les styles aux cellules
        for row_idx in range(2, ws1.max_row + 1):
            for col_idx in range(1, ws1.max_column + 1):
                cell = ws1.cell(row_idx, col_idx)
                cell.border = thin_border

                # Lignes TOTAL PVT en vert clair et TOTAL en orange clair
                if cell.value == 'TOTAL PVT':
                    for col in range(1, ws1.max_column + 1):
                        total_cell = ws1.cell(row_idx, col)
                        total_cell.font = Font(bold=True, size=11)
                        total_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        total_cell.alignment = Alignment(horizontal='center', vertical='center')
                elif cell.value == 'TOTAL':
                    for col in range(1, ws1.max_column + 1):
                        total_cell = ws1.cell(row_idx, col)
                        total_cell.font = Font(bold=True, size=11)
                        total_cell.fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
                        total_cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(size=10)

        # Largeurs des colonnes
        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 45
        ws1.column_dimensions['C'].width = 18
        ws1.column_dimensions['D'].width = 18
        ws1.column_dimensions['E'].width = 20
        ws1.freeze_panes = 'A2'

        # Formater "Résumé PVT"
        ws2 = wb['Résumé PVT']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Fusionner DRV pour Résumé PVT
        drv_ranges_pvt = []
        current_drv = None
        drv_start = 2

        for row_idx in range(2, ws2.max_row):
            drv_value = ws2.cell(row_idx, 1).value
            if drv_value and drv_value != current_drv:
                if current_drv is not None and row_idx > drv_start:
                    drv_ranges_pvt.append((drv_start, row_idx - 1, current_drv))
                current_drv = drv_value
                drv_start = row_idx

        if ws2.max_row - 1 >= drv_start:
            drv_ranges_pvt.append((drv_start, ws2.max_row - 1, current_drv))

        for start_row, end_row, drv_value in drv_ranges_pvt:
            if end_row > start_row:
                ws2.merge_cells(f'A{start_row}:A{end_row}')
                ws2.cell(start_row, 1).alignment = Alignment(horizontal='left', vertical='center')
                ws2.cell(start_row, 1).font = Font(bold=True, size=10)

        # Appliquer les styles
        for row_idx in range(2, ws2.max_row + 1):
            for col_idx in range(1, ws2.max_column + 1):
                cell = ws2.cell(row_idx, col_idx)
                cell.border = thin_border

                # Ligne TOTAL
                if row_idx == ws2.max_row:
                    cell.font = Font(bold=True, size=11)
                    cell.fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    if col_idx in [1, 2, 3]:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(size=10)

        # Fusionner colonnes pour ligne TOTAL
        ws2.merge_cells(f'A{ws2.max_row}:C{ws2.max_row}')
        ws2.cell(ws2.max_row, 1).value = 'TOTAL GÉNÉRAL'

        # Largeurs des colonnes
        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 45
        ws2.column_dimensions['C'].width = 20
        ws2.column_dimensions['D'].width = 18
        ws2.column_dimensions['E'].width = 18
        ws2.column_dimensions['F'].width = 18
        ws2.freeze_panes = 'A2'

        # Sauvegarder
        final_buffer = BytesIO()
        wb.save(final_buffer)
        final_buffer.seek(0)

        st.download_button(
            label="📥 Télécharger le Rapport Paiement Mensuel (Excel Formaté)",
            data=final_buffer,
            file_name="paiement_mensuel_global.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"❌ Erreur lors de la génération du fichier Excel : {str(e)}")
        import traceback
        st.code(traceback.format_exc())

# ====================
# Footer
# ====================
st.markdown("""
<div style="
    margin-top: 2rem;
    padding: 1rem;
    background: linear-gradient(135deg, #009CA6 0%, #00B8C5 100%);
    border-radius: 12px;
    text-align: center;
    color: white;
    box-shadow: 0 4px 15px rgba(0, 156, 166, 0.3);
">
    <p style="margin: 0; font-size: 0.9rem; font-weight: 500;">
        Propulsé par Orange Sénégal - Sonatel SA | 2025
    </p>
</div>
""", unsafe_allow_html=True)