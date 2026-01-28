import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

# ====================
# Configuration page
# ====================
st.set_page_config(page_title="LOUMA - Reporting Multi-Mois", layout="wide")

# ====================
# COULEURS ET MAPPING
# ====================
SONATEL_COLORS = {
    "orange_primary": "FFFF7900",
    "orange_dark": "FFFF5000",
    "light_orange": "FFFFE5CC",
    "blue_dark": "FF003366",
    "white": "FFFFFFFF",
    "gray_light": "FFF5F5F5"
}

DRV_MAPPING = {
    "DV-DRV2_DIRECTION REGIONALE DES VENTES DAKAR 2": "DR2",
    "DV-DRVS_DIRECTION REGIONALE DES VENTES SUD": "DRS",
    "DV-DRVSE_DIRECTION REGIONALE DES VENTES SUD-EST": "DRSE",
    "DV-DRVN_DIRECTION REGIONALE DES VENTES NORD": "DRN",
    "DV-DRVC_DIRECTION REGIONALE DES VENTES CENTRE": "DRC",
    "DV-DRVE_DIRECTION REGIONALE DES VENTES EST": "DRE"
}

# ❌ LISTE DES EXCLUSIONS DEMANDÉES
EXCLUSIONS = [
    "DMGP-KMOB_AGENCE KIRENE MOBILE",
    "DMGP-DMC_DEPARTEMENT MARKETING DE LA COMMERCIALISATION",
    "DESC-DIS_DEPARTEMENT INGENIERIE DU SERVICE",
    "DDE-DRC_DEPARTEMENT RELATION CLIENT",
    "DAL-CGL_DEPARTEMENT GESTION DE LA CHAINE LOGISTIQUE",
    "DV-DVRI_DEVELOPPEMENT DES VENTES DU RESEAU INDIRECT"
]

# CSS pour le look Sonatel
st.markdown(f"""
<style>
    .section-title {{ background: linear-gradient(135deg, #FF5000 0%, #FF7900 100%); color: white; padding: 0.8rem; border-radius: 10px; font-weight: 600; text-align: center; margin-bottom: 1.2rem; }}
    .metric-card {{ background: white; border-radius: 12px; padding: 1.5rem; box-shadow: 0 4px 15px rgba(255, 121, 0, 0.15); border: 2px solid #FFE5CC; text-align: center; }}
    .metric-value {{ font-size: 2.2rem; font-weight: 700; color: #FF7900; }}
</style>
""", unsafe_allow_html=True)

# ====================
# FONCTIONS DE TRAITEMENT
# ====================

def load_data_file(uploaded_file):
    try:
        if uploaded_file.name.endswith('.csv'):
            df = pd.read_csv(uploaded_file, encoding='utf-8', sep=';')
        else:
            df = pd.read_excel(uploaded_file)
        df.columns = df.columns.str.strip()
        return df, True
    except:
        return pd.DataFrame(), False

def process_single_month(df, details, mois_nom):
    mapping = {
        'MSISDN': 'REALISATION',
        'ACCUEIL_VENDEUR': 'PVT',
        'AGENCE_VENDEUR': 'DR_ORIGINE', # On garde temporairement le nom d'origine pour filtrer
        'ETAT_IDENTIFICATION': 'ETAT'
    }
    df = df.rename(columns=mapping)

    # 1. Filtre sur l'état d'identification
    if 'ETAT' in df.columns:
        df = df[df['ETAT'].astype(str).str.strip().isin(details)]

    # 2. FILTRE D'EXCLUSION DES AGENCES/DÉPARTEMENTS
    if 'DR_ORIGINE' in df.columns:
        # On ne garde que ce qui n'est PAS dans la liste EXCLUSIONS
        df = df[~df['DR_ORIGINE'].isin(EXCLUSIONS)]

        # Ensuite on applique le mapping (DR2, DRS, etc.)
        df["DR"] = df["DR_ORIGINE"].replace(DRV_MAPPING)
        # Si une agence n'est pas dans le mapping Sonatel et n'est pas exclue, on peut la nommer "AUTRE"
        df.loc[~df["DR_ORIGINE"].isin(DRV_MAPPING.keys()), "DR"] = "HORS DR"

    df['MOIS_NOM'] = mois_nom
    return df

def process_multi_month_data(df_list, mois_noms, details, obj_mensuel=960):
    all_dfs = [process_single_month(df, details, m) for df, m in zip(df_list, mois_noms) if not df.empty]
    if not all_dfs: return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    df_combined = pd.concat(all_dfs, ignore_index=True)
    pvt_counts = df_combined.groupby(['DR', 'PVT'])['MOIS_NOM'].nunique().reset_index(name='NB_MOIS')

    df_pvt = df_combined.groupby(['DR', 'PVT'], as_index=False).size().rename(columns={'size': 'REALISATION'})
    df_pvt = pd.merge(df_pvt, pvt_counts, on=['DR', 'PVT'])
    df_pvt['OBJECTIF'] = df_pvt['NB_MOIS'] * obj_mensuel
    df_pvt['R/O'] = (df_pvt['REALISATION'] / df_pvt['OBJECTIF'] * 100).fillna(0).astype(int)

    df_dr = df_pvt.groupby('DR', as_index=False).agg({'REALISATION': 'sum', 'OBJECTIF': 'sum'})
    df_dr['R/O'] = (df_dr['REALISATION'] / df_dr['OBJECTIF'] * 100).fillna(0).astype(int)

    return df_combined, df_pvt, df_dr

def generate_beautiful_excel(df_dr, df_pvt, annee):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_dr.to_excel(writer, sheet_name='Résumé DR', index=False)
        df_pvt.to_excel(writer, sheet_name='Résumé PVT', index=False)

    wb = load_workbook(output)
    header_fill = PatternFill(start_color=SONATEL_COLORS['orange_primary'], end_color=SONATEL_COLORS['orange_primary'], fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal='center')

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                header_val = str(ws.cell(row=1, column=cell.column).value)
                if "R/O" in header_val:
                    val = cell.value if isinstance(cell.value, (int, float)) else 0
                    if val >= 100: cell.font = Font(color="00B050", bold=True)
                    elif val >= 80: cell.font = Font(color="FF7900", bold=True)
                    else: cell.font = Font(color="FF0000", bold=True)
                    cell.value = f"{int(val)}%"

    output.seek(0)
    wb.save(output)
    return output.getvalue()

# ====================
# APPLICATION
# ====================
def main():
    st.markdown('<div class="section-title">📅 CONFIGURATION DE LA PÉRIODE</div>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    nb_mois = c1.selectbox("Nombre de mois", list(range(1, 13)), index=10)
    annee = c2.selectbox("Année", [2024, 2025, 2026], index=1)

    mois_options = ['Janvier','Février','Mars','Avril','Mai','Juin','Juillet','Août','Septembre','Octobre','Novembre','Décembre']
    uploaded_files = []
    cols = st.columns(3)

    for i in range(nb_mois):
        with cols[i % 3]:
            m_nom = st.selectbox(f"Mois {i+1}", mois_options, key=f"m{i}", index=i)
            f = st.file_uploader(f"Fichier {m_nom}", type=["xlsx", "csv"], key=f"f{i}")
            if f: uploaded_files.append((m_nom, f))

    if st.button("🚀 GÉNÉRER LE RAPPORT") and uploaded_files:
        with st.spinner("Analyse en cours..."):
            details = ["En Cours-Identification", "Identifie", "Identifie Photo"]
            df_list, noms = [], []
            for m, f in uploaded_files:
                df, ok = load_data_file(f)
                if ok:
                    df_list.append(df)
                    noms.append(m)

            df_c, df_p, df_dr = process_multi_month_data(df_list, noms, details)

            if not df_dr.empty:
                st.success("✅ Analyse terminée (Exclusions appliquées) !")
                real, obj = df_dr['REALISATION'].sum(), df_dr['OBJECTIF'].sum()

                k1, k2, k3 = st.columns(3)
                k1.markdown(f'<div class="metric-card"><div class="metric-value">{real:,}</div>Ventes</div>', unsafe_allow_html=True)
                k2.markdown(f'<div class="metric-card"><div class="metric-value">{obj:,}</div>Objectif</div>', unsafe_allow_html=True)
                k3.markdown(f'<div class="metric-card"><div class="metric-value">{(real/obj*100):.1f}%</div>R/O Global</div>', unsafe_allow_html=True)

                st.subheader("Performance par DR")
                st.dataframe(df_dr.style.format({'REALISATION': '{:,}', 'OBJECTIF': '{:,}', 'R/O': '{}%'}))

                excel_data = generate_beautiful_excel(df_dr, df_p, annee)
                st.download_button("📥 Télécharger le rapport Excel", excel_data, f"Louma_{annee}.xlsx", "application/vnd.openxml")

if __name__ == "__main__":
    main()