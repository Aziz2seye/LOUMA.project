# reporting_mensuel_louma.py
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import sys
from pathlib import Path
from PIL import Image
import plotly.express as px
import plotly.graph_objects as go
import plotly.io as pio
from datetime import datetime
import calendar

# Configuration pour l'export PNG
pio.kaleido.scope.default_format = "png"
pio.kaleido.scope.default_width = 800
pio.kaleido.scope.default_height = 400
pio.kaleido.scope.default_scale = 1.5

# Ajouter le répertoire parent au path Python
current_dir = Path(__file__).parent
parent_dir = current_dir.parent
sys.path.insert(0, str(parent_dir))

# ====================
# FONCTION D'EXPORT DES GRAPHIQUES EN PNG
# ====================
def download_plotly_as_png(fig, filename):
    """
    Convertit un graphique Plotly en PNG téléchargeable
    """
    import io
    buffer = io.BytesIO()
    fig.write_image(buffer, format='png')
    buffer.seek(0)
    return buffer

# ====================
# Configuration page
# ====================
st.set_page_config(page_title="LOUMA - Reporting Mensuel", layout="wide", initial_sidebar_state="expanded")

# ====================
# CSS personnalisé Orange Sonatel
# ====================
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600;700&display=swap');

    header[data-testid="stHeader"] { display: none; }

    .block-container {
        padding-top: 2rem !important;
        padding-bottom: 2rem !important;
    }

    .main {
        font-family: 'Poppins', sans-serif;
        background: linear-gradient(135deg, #fff5f0 0%, #ffffff 50%, #f0f8ff 100%);
    }

    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #FF7900 0%, #FF5000 100%) !important;
    }

    section[data-testid="stSidebar"] * {
        color: white !important;
    }

    .stDataFrame {
        border-radius: 15px;
        overflow: hidden;
        box-shadow: 0 8px 20px rgba(255, 121, 0, 0.15);
        border: 2px solid #FFE5CC;
    }

    .section-title {
        background: linear-gradient(135deg, #FF7900 0%, #FF5000 100%);
        color: white;
        padding: 0.7rem 1.2rem;
        border-radius: 10px;
        font-weight: 600;
        font-size: 1.1rem;
        margin-bottom: 0.8rem;
        box-shadow: 0 4px 12px rgba(255, 121, 0, 0.25);
        text-align: center;
        max-width: 500px;
        margin-left: auto;
        margin-right: auto;
    }

    .stButton > button {
        background: linear-gradient(135deg, #FF7900 0%, #FF5000 100%);
        color: white;
        border: none;
        border-radius: 10px;
        padding: 0.8rem 2rem;
        font-weight: 600;
        font-size: 1.1rem;
        box-shadow: 0 4px 12px rgba(255, 121, 0, 0.3);
        transition: all 0.3s ease;
        width: 100%;
    }

    .stButton > button:hover {
        background: linear-gradient(135deg, #FF5000 0%, #FF3000 100%);
        box-shadow: 0 6px 18px rgba(255, 121, 0, 0.5);
        transform: translateY(-2px);
    }

    .stDownloadButton > button {
        background: linear-gradient(135deg, #00D4AA 0%, #00B890 100%);
        color: white;
        border: none;
        border-radius: 10px;
        padding: 0.8rem 2rem;
        font-weight: 600;
        font-size: 1.1rem;
        box-shadow: 0 4px 12px rgba(0, 212, 170, 0.3);
        transition: all 0.3s ease;
        width: 100%;
    }

    .stDownloadButton > button:hover {
        background: linear-gradient(135deg, #00B890 0%, #009A7A 100%);
        boxshadow: 0 6px 18px rgba(0, 212, 170, 0.5);
        transform: translateY(-2px);
    }

    .metric-card {
        background: white;
        border-radius: 12px;
        padding: 1.5rem;
        box-shadow: 0 4px 15px rgba(255, 121, 0, 0.15);
        border: 2px solid #FFE5CC;
        text-align: center;
    }

    .metric-value {
        font-size: 2rem;
        font-weight: 700;
        color: #FF7900;
    }

    .metric-label {
        font-size: 1rem;
        color: #666;
        margin-top: 0.5rem;
    }
</style>
""", unsafe_allow_html=True)

# ====================
# Charger le logo
# ====================
logo = None
logo_paths = [
    parent_dir / "assets" / "logo sonatel.png",
    Path("assets") / "logo sonatel.png",
    Path("../assets/logo sonatel.png"),
]

for logo_path in logo_paths:
    try:
        if logo_path.exists():
            logo = Image.open(logo_path)
            break
    except:
        continue

# ====================
# Header avec logo et titre
# ====================
col_logo, col_title = st.columns([1, 3])

with col_logo:
    if logo:
        st.image(logo, width=280)

with col_title:
    st.markdown("""
    <div style="
        background: linear-gradient(135deg, #FF7900 0%, #FF5000 100%);
        padding: 2rem;
        border-radius: 20px;
        box-shadow: 0 8px 25px rgba(255, 121, 0, 0.4);
        display: flex;
        flex-direction: column;
        justify-content: center;
        border: 3px solid rgba(255, 255, 255, 0.2);
        height: 100%;
    ">
        <h1 style="
            color: white;
            font-size: 2.5rem;
            font-weight: 700;
            margin: 0;
            text-shadow: 3px 3px 10px rgba(0, 0, 0, 0.3);
        ">
            📈 Reporting Mensuel
        </h1>
        <p style="
            color: rgba(255, 255, 255, 0.95);
            font-size: 1.2rem;
            margin: 0.8rem 0 0 0;
            font-weight: 400;
        ">
            Analyse des performances commerciales - Orange Sénégal
        </p>
    </div>
    """, unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# 🎯 Mapping DRV unique
DRV_MAPPING = {
    "DV-DRV2_DIRECTION REGIONALE DES VENTES DAKAR 2": "DR2",
    "DV-DRVS_DIRECTION REGIONALE DES VENTES SUD": "DRS",
    "DV-DRVSE_DIRECTION REGIONALE DES VENTES SUD-EST": "DRSE",
    "DV-DRVN_DIRECTION REGIONALE DES VENTES NORD": "DRN",
    "DV-DRVC_DIRECTION REGIONALE DES VENTES CENTRE": "DRC",
    "DV-DRVE_DIRECTION REGIONALE DES VENTES EST": "DRE"
}

# ====================
# FONCTION POUR LE CHARGEMENT DES DONNÉES VTO
# ====================
def load_vto():
    """Fonction pour charger les données VTO"""
    try:
        from utils import load_vto as real_load_vto
        return real_load_vto()
    except:
        data = {
            'LOGIN': ['vto001', 'vto002', 'vto003', 'vto004', 'vto005', 'vto006', 'vto007', 'vto008'],
            'NOM': ['DIOUF', 'NDIAYE', 'SARR', 'FALL', 'DIOP', 'GUEYE', 'MBOW', 'SY'],
            'PRENOM': ['Mamadou', 'Fatou', 'Moussa', 'Aïssatou', 'Cheikh', 'Alioune', 'Mariama', 'Oumar'],
            'DR': ['DR2', 'DRS', 'DRSE', 'DRN', 'DRC', 'DRE', 'DR2', 'DRS'],
            'PVT': ['PVT DAKAR CENTRE', 'PVT SUD', 'PVT SUD-EST', 'PVT NORD', 'PVT CENTRE', 'PVT EST', 'PVT DAKAR PLATEAU', 'PVT SUD 2']
        }
        return pd.DataFrame(data)

# ====================
# FONCTIONS DE TRAITEMENT DES DONNÉES
# ====================
def process_monthly_data(df, vto_df, details, objectif_pvt=960):
    """Fonction pour traiter les données mensuelles"""
    column_mapping = {
        'MSISDN': 'REALISATION',
        'ACCUEIL_VENDEUR': 'PVT',
        'LOGIN_VENDEUR': 'LOGIN',
        'AGENCE_VENDEUR': 'DR',
        'NOM_VENDEUR': 'NOM_VENDEUR',
        'PRENOM_VENDEUR': 'PRENOM_VENDEUR',
        'ETAT_IDENTIFICATION': 'ETAT_IDENTIFICATION'
    }

    for original_col, new_col in column_mapping.items():
        if original_col in df.columns:
            df = df.rename(columns={original_col: new_col})

    if 'LOGIN' in df.columns:
        df['LOGIN'] = df['LOGIN'].astype(str).str.lower().str.strip()

    if 'DR' in df.columns:
        df['DR'] = df['DR'].astype(str).str.strip().str.upper()

    if 'NOM_VENDEUR' in df.columns:
        df['NOM_VENDEUR'] = df['NOM_VENDEUR'].astype(str).str.strip().str.upper()

    if 'PRENOM_VENDEUR' in df.columns:
        df['PRENOM_VENDEUR'] = df['PRENOM_VENDEUR'].astype(str).str.strip().str.upper()

    logins_concernes = vto_df["LOGIN"].astype(str).str.lower().str.strip().tolist()

    df_filtre_login = df[df['LOGIN'].isin(logins_concernes)] if 'LOGIN' in df.columns else df

    if 'ETAT_IDENTIFICATION' in df_filtre_login.columns:
        df_filtre = df_filtre_login[df_filtre_login['ETAT_IDENTIFICATION'].astype(str).isin(details)]
    else:
        df_filtre = df_filtre_login

    if 'DR' in df_filtre.columns:
        df_filtre["DR"] = df_filtre["DR"].replace(DRV_MAPPING)
        df_filtre["DR"] = df_filtre["DR"].apply(lambda x: x if x in DRV_MAPPING.values() else 'AUTRE')

    # 1. Résumé par PVT
    if all(col in df_filtre.columns for col in ['DR', 'PVT']):
        df_pvt_summary = df_filtre.groupby(['DR', 'PVT'], as_index=False).size()
        df_pvt_summary.columns = ['DR', 'PVT', 'REALISATION']
        df_pvt_summary['OBJECTIF'] = objectif_pvt
        df_pvt_summary['R/O'] = (df_pvt_summary['REALISATION'] / df_pvt_summary['OBJECTIF'] * 100).round(0).astype(int)
        df_pvt_summary = df_pvt_summary.sort_values(['DR', 'PVT'])
    else:
        df_pvt_summary = pd.DataFrame(columns=['DR', 'PVT', 'REALISATION', 'OBJECTIF', 'R/O'])

    # 2. Résumé par DR
    if 'DR' in df_filtre.columns:
        df_dr_summary = df_filtre.groupby('DR', as_index=False).size()
        df_dr_summary.columns = ['DR', 'REALISATION']

        if not df_pvt_summary.empty:
            pvt_count_per_dr = df_pvt_summary.groupby('DR')['PVT'].nunique()
            df_dr_summary['OBJECTIF'] = df_dr_summary['DR'].map(pvt_count_per_dr) * objectif_pvt
        else:
            df_dr_summary['OBJECTIF'] = 0

        df_dr_summary['R/O'] = (df_dr_summary['REALISATION'] / df_dr_summary['OBJECTIF'] * 100).round(0).astype(int)
        df_dr_summary = df_dr_summary.sort_values('DR')
    else:
        df_dr_summary = pd.DataFrame(columns=['DR', 'REALISATION', 'OBJECTIF', 'R/O'])

    # 3. Détails par VTO
    required_vto_cols = ['DR', 'PVT', 'PRENOM_VENDEUR', 'NOM_VENDEUR', 'LOGIN']
    if all(col in df_filtre.columns for col in required_vto_cols):
        df_reporting = df_filtre.groupby(['DR', 'PVT', 'PRENOM_VENDEUR', 'NOM_VENDEUR', 'LOGIN'], as_index=False).size()
        df_reporting.columns = ['DR', 'PVT', 'PRENOM_VENDEUR', 'NOM_VENDEUR', 'LOGIN', 'REALISATION']
        df_reporting = df_reporting.sort_values(['DR', 'PVT', 'REALISATION'], ascending=[True, True, False])
    else:
        df_reporting = pd.DataFrame(columns=['DR', 'PVT', 'PRENOM_VENDEUR', 'NOM_VENDEUR', 'LOGIN', 'REALISATION'])

    return df_filtre, df_pvt_summary, df_dr_summary, df_reporting

def generate_monthly_excel_report(df_pvt_summary, df_dr_summary, df_reporting, mois_nom, annee):
    """Génère un fichier Excel avec mise en forme pour le reporting mensuel"""
    buffer_output = BytesIO()

    # Totaux pour PVT
    total_realisation_pvt = int(df_pvt_summary['REALISATION'].sum())
    total_objectif_pvt = int(df_pvt_summary['OBJECTIF'].sum())
    total_ro_pvt = round((total_realisation_pvt / total_objectif_pvt * 100), 0) if total_objectif_pvt > 0 else 0

    # Totaux pour DR
    total_realisation_dr = int(df_dr_summary['REALISATION'].sum())
    total_objectif_dr = int(df_dr_summary['OBJECTIF'].sum())
    total_ro_dr = round((total_realisation_dr / total_objectif_dr * 100), 0) if total_objectif_dr > 0 else 0

    # Préparer le DataFrame pour l'export DR
    df_dr_summary_export = df_dr_summary.copy().reset_index(drop=True)
    new_row_dr = {
        'DR': 'TOTAL',
        'REALISATION': total_realisation_dr,
        'OBJECTIF': total_objectif_dr,
        'R/O': int(total_ro_dr)
    }
    df_dr_summary_export = pd.concat([
        df_dr_summary_export,
        pd.DataFrame([new_row_dr])
    ], ignore_index=True)

    # Préparer le DataFrame pour l'export PVT
    df_pvt_summary_export = df_pvt_summary.copy().reset_index(drop=True)
    new_row_pvt = {
        'DR': '',
        'PVT': 'TOTAL',
        'REALISATION': total_realisation_pvt,
        'OBJECTIF': total_objectif_pvt,
        'R/O': int(total_ro_pvt)
    }
    df_pvt_summary_export = pd.concat([
        df_pvt_summary_export,
        pd.DataFrame([new_row_pvt])
    ], ignore_index=True)

    # Préparer le DataFrame pour l'export VTO
    total_realisation_vto = int(df_reporting['REALISATION'].sum())
    df_reporting_export = df_reporting.copy().reset_index(drop=True)
    new_row_vto = {
        'DR': '',
        'PVT': '',
        'PRENOM_VENDEUR': '',
        'NOM_VENDEUR': '',
        'LOGIN': 'TOTAL',
        'REALISATION': total_realisation_vto
    }
    df_reporting_export = pd.concat([
        df_reporting_export,
        pd.DataFrame([new_row_vto])
    ], ignore_index=True)

    # Écrire dans l'ordre demandé : Résumé DR, Résumé PVT, Détails VTO
    with pd.ExcelWriter(buffer_output, engine='openpyxl') as writer:
        df_dr_summary_export.to_excel(writer, sheet_name='Résumé DR', index=False)
        df_pvt_summary_export.to_excel(writer, sheet_name='Résumé PVT', index=False)
        df_reporting_export.to_excel(writer, sheet_name='Détails VTO', index=False)

    buffer_output.seek(0)
    wb = load_workbook(buffer_output)

    # ============================================
    # COULEURS POUR LA COLORATION CONDITIONNELLE
    # ============================================
    # Vert pour R/O >= 100%
    vert_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    vert_font = Font(color="006100")

    # Rouge pour R/O < 80%
    rouge_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    rouge_font = Font(color="9C0006")

    # Jaune pour R/O entre 80 et 99%
    jaune_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    jaune_font = Font(color="9C6500")

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ============================================
    # Formater la feuille "Résumé DR"
    # ============================================
    ws_dr = wb['Résumé DR']

    # Formater les en-têtes (ligne 1)
    for cell in ws_dr[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Mettre à jour les en-têtes
    ws_dr['B1'].value = 'REALISATION'
    ws_dr['C1'].value = 'OBJECTIF'
    ws_dr['D1'].value = 'R/O (%)'

    # Formater les données
    for row_idx in range(2, ws_dr.max_row + 1):
        for col_idx in range(1, 5):
            cell = ws_dr.cell(row_idx, col_idx)
            cell.border = thin_border

            # Appliquer le formatage spécial pour la colonne R/O (colonne D)
            if col_idx == 4:  # Colonne R/O
                ro_value = cell.value
                if ro_value is not None:
                    try:
                        ro_numeric = float(ro_value)

                        # Appliquer le symbole % directement dans la cellule
                        cell.value = f"{int(ro_numeric)}%"

                        # Appliquer la couleur selon la valeur
                        if ro_numeric >= 100:
                            cell.fill = vert_fill
                            cell.font = vert_font
                        elif 80 <= ro_numeric <= 99:
                            cell.fill = jaune_fill
                            cell.font = jaune_font
                        elif ro_numeric < 80:
                            cell.fill = rouge_fill
                            cell.font = rouge_font
                    except (ValueError, AttributeError):
                        pass

            if row_idx == ws_dr.max_row:  # Ligne TOTAL
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
                if col_idx in [1, 2, 3, 4]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                if col_idx == 1:  # Colonne DR
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.font = Font(bold=True, size=10)
                elif col_idx in [2, 3]:  # Colonnes REALISATION et OBJECTIF
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(size=10)
                elif col_idx == 4:  # Colonne R/O
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(bold=True, size=10)

    ws_dr.column_dimensions['A'].width = 15
    ws_dr.column_dimensions['B'].width = 12
    ws_dr.column_dimensions['C'].width = 12
    ws_dr.column_dimensions['D'].width = 10
    ws_dr.freeze_panes = 'A2'

    # ============================================
    # Formater la feuille "Résumé PVT"
    # ============================================
    ws_pvt = wb['Résumé PVT']

    # Formater les en-têtes (ligne 1)
    for cell in ws_pvt[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Mettre à jour les en-têtes
    ws_pvt['C1'].value = 'REALISATION'
    ws_pvt['E1'].value = 'R/O (%)'

    # Fusionner les cellules pour les DR
    drv_ranges_pvt = []
    current_drv = None
    drv_start = 2

    for row_idx in range(2, ws_pvt.max_row):
        drv_value = ws_pvt.cell(row_idx, 1).value
        if drv_value and drv_value != current_drv:
            if current_drv is not None and row_idx > drv_start:
                drv_ranges_pvt.append((drv_start, row_idx - 1, current_drv))
            current_drv = drv_value
            drv_start = row_idx

    if ws_pvt.max_row - 1 >= drv_start:
        drv_ranges_pvt.append((drv_start, ws_pvt.max_row - 1, current_drv))

    for start_row, end_row, drv_value in drv_ranges_pvt:
        if end_row > start_row:
            ws_pvt.merge_cells(f'A{start_row}:A{end_row}')
            ws_pvt.cell(start_row, 1).alignment = Alignment(horizontal='left', vertical='center')
            ws_pvt.cell(start_row, 1).font = Font(bold=True, size=10)

    # Fusionner les cellules pour le TOTAL
    ws_pvt.merge_cells(f'A{ws_pvt.max_row}:B{ws_pvt.max_row}')
    ws_pvt.cell(ws_pvt.max_row, 1).value = 'TOTAL'

    # Appliquer la coloration conditionnelle
    for row_idx in range(2, ws_pvt.max_row + 1):
        for col_idx in range(1, 6):
            cell = ws_pvt.cell(row_idx, col_idx)
            cell.border = thin_border

            # Appliquer le formatage spécial pour la colonne R/O (colonne E)
            if col_idx == 5 and row_idx < ws_pvt.max_row:  # Colonne R/O, sauf la ligne TOTAL
                ro_value = cell.value
                if ro_value is not None:
                    try:
                        ro_numeric = float(ro_value)

                        # Appliquer la couleur selon la valeur
                        if ro_numeric >= 100:
                            cell.fill = vert_fill
                            cell.font = vert_font
                            cell.value = f"{int(ro_numeric)}%"
                        elif 80 <= ro_numeric <= 99:
                            cell.fill = jaune_fill
                            cell.font = jaune_font
                            cell.value = f"{int(ro_numeric)}%"
                        elif ro_numeric < 80:
                            cell.fill = rouge_fill
                            cell.font = rouge_font
                            cell.value = f"{int(ro_numeric)}%"
                    except (ValueError, AttributeError):
                        pass

            if row_idx == ws_pvt.max_row:  # Ligne TOTAL
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
                if col_idx == 5:  # Colonne R/O pour le TOTAL
                    cell.value = f"{int(cell.value)}%" if cell.value else "0%"
                if col_idx in [1, 3, 4, 5]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx in [1, 2]:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.font = Font(size=10)
            elif col_idx in [3, 4]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(size=10)
            elif col_idx == 5:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True, size=10)

    ws_pvt.column_dimensions['A'].width = 8
    ws_pvt.column_dimensions['B'].width = 45
    ws_pvt.column_dimensions['C'].width = 12
    ws_pvt.column_dimensions['D'].width = 12
    ws_pvt.column_dimensions['E'].width = 10
    ws_pvt.freeze_panes = 'A2'

    # ============================================
    # Formater la feuille "Détails VTO"
    # ============================================
    ws = wb['Détails VTO']

    # Formater les en-têtes (ligne 1)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Mettre à jour les en-têtes
    ws['F1'].value = 'REALISATION'

    # Fusionner les cellules pour les DR
    drv_ranges = []
    current_drv = None
    drv_start = 2

    for row_idx in range(2, ws.max_row):
        drv_value = ws.cell(row_idx, 1).value
        if drv_value != current_drv:
            if current_drv is not None and row_idx > drv_start:
                drv_ranges.append((drv_start, row_idx - 1, current_drv))
            current_drv = drv_value
            drv_start = row_idx

    if ws.max_row - 1 >= drv_start:
        drv_ranges.append((drv_start, ws.max_row - 1, current_drv))

    # Fusionner les cellules pour les PVT
    pvt_ranges = []
    current_pvt = None
    pvt_start = 2

    for row_idx in range(2, ws.max_row):
        pvt_value = ws.cell(row_idx, 2).value
        if pvt_value != current_pvt:
            if current_pvt is not None and row_idx > pvt_start:
                pvt_ranges.append((pvt_start, row_idx - 1, current_pvt))
            current_pvt = pvt_value
            pvt_start = row_idx

    if ws.max_row - 1 >= pvt_start:
        pvt_ranges.append((pvt_start, ws.max_row - 1, current_pvt))

    # Appliquer les fusions pour les DR
    for start_row, end_row, drv_value in drv_ranges:
        if end_row > start_row:
            ws.merge_cells(f'A{start_row}:A{end_row}')
            ws.cell(start_row, 1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(start_row, 1).font = Font(bold=True, size=10)

    # Appliquer les fusions pour les PVT
    for start_row, end_row, pvt_value in pvt_ranges:
        if end_row > start_row:
            ws.merge_cells(f'B{start_row}:B{end_row}')
            ws.cell(start_row, 2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws.cell(start_row, 2).font = Font(bold=True, size=10)

    # Fusionner pour le TOTAL
    ws.merge_cells(f'A{ws.max_row}:E{ws.max_row}')
    ws.cell(ws.max_row, 1).value = 'TOTAL'

    # Formater toutes les cellules
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, 7):
            cell = ws.cell(row_idx, col_idx)
            cell.border = thin_border

            if row_idx == ws.max_row:  # Ligne TOTAL
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
                if col_idx in [1, 6]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx in [1, 2, 3, 4, 5]:
                if col_idx not in [1, 2]:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.font = Font(size=10)
            elif col_idx == 6:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(size=10)

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 12
    ws.freeze_panes = 'A2'

    final_buffer = BytesIO()
    wb.save(final_buffer)
    final_buffer.seek(0)

    return final_buffer

def display_monthly_metrics(df_filtre, df_reporting, df_pvt_summary, df_dr_summary, mois_nom, annee):
    """Affiche les métriques et graphiques pour le reporting mensuel"""

    st.markdown('<div class="section-title">📊 MÉTRIQUES PRINCIPALES</div>', unsafe_allow_html=True)

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        total_ventes = df_filtre.shape[0]
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{total_ventes:,}</div>
            <div class="metric-label">Ventes Totales</div>
        </div>
        """, unsafe_allow_html=True)

    with col2:
        vto_actifs = df_filtre['LOGIN'].nunique() if 'LOGIN' in df_filtre.columns else 0
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{vto_actifs}</div>
            <div class="metric-label">VTO Actifs</div>
        </div>
        """, unsafe_allow_html=True)

    with col3:
        pvt_concernes = df_filtre['PVT'].nunique() if 'PVT' in df_filtre.columns else 0
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{pvt_concernes}</div>
            <div class="metric-label">PVT Concernés</div>
        </div>
        """, unsafe_allow_html=True)

    with col4:
        taux_realisation = round((total_ventes / (df_pvt_summary['OBJECTIF'].sum() if not df_pvt_summary.empty else 1) * 100), 1)
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{taux_realisation}%</div>
            <div class="metric-label">Taux de Réalisation</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # 📊 TABLEAU 1 : Résumé par DR
    st.markdown(f'<div class="section-title">📋 RÉSUMÉ PAR DIRECTION RÉGIONALE - {mois_nom.upper()} {annee}</div>', unsafe_allow_html=True)

    if not df_dr_summary.empty:
        df_dr_summary_display = df_dr_summary.copy()
        df_dr_summary_display['R/O'] = df_dr_summary_display['R/O'].astype(str) + '%'

        total_realisation = df_dr_summary['REALISATION'].sum()
        total_objectif = df_dr_summary['OBJECTIF'].sum()
        total_ro = round((total_realisation / total_objectif * 100), 0) if total_objectif > 0 else 0

        df_dr_summary_display.loc[len(df_dr_summary_display)] = ['TOTAL', total_realisation, total_objectif, f'{int(total_ro)}%']

        def color_ro_dr(val):
            if isinstance(val, str) and val.endswith('%'):
                try:
                    ro_numeric = float(val.strip('%'))
                    if ro_numeric >= 100:
                        return 'background-color: #C6EFCE; color: #006100; font-weight: bold;'
                    elif 80 <= ro_numeric <= 99:
                        return 'background-color: #FFEB9C; color: #9C6500; font-weight: bold;'
                    elif ro_numeric < 80:
                        return 'background-color: #FFC7CE; color: #9C0006; font-weight: bold;'
                except:
                    pass
            return ''

        styled_df_dr = df_dr_summary_display.style.applymap(color_ro_dr, subset=['R/O'])
        st.dataframe(styled_df_dr, use_container_width=True)
    else:
        st.warning("Aucune donnée disponible pour le résumé par DR")

    st.markdown("<br>", unsafe_allow_html=True)

    # 📊 TABLEAU 2 : Résumé par PVT
    st.markdown(f'<div class="section-title">📋 RÉSUMÉ PAR POINT DE VENTE - {mois_nom.upper()} {annee}</div>', unsafe_allow_html=True)

    if not df_pvt_summary.empty:
        df_pvt_summary_display = df_pvt_summary.copy()
        df_pvt_summary_display['R/O'] = df_pvt_summary_display['R/O'].astype(str) + '%'

        total_realisation = df_pvt_summary['REALISATION'].sum()
        total_objectif = df_pvt_summary['OBJECTIF'].sum()
        total_ro = round((total_realisation / total_objectif * 100), 0) if total_objectif > 0 else 0

        df_pvt_summary_display.loc[len(df_pvt_summary_display)] = ['', 'TOTAL', total_realisation, total_objectif, f'{int(total_ro)}%']

        def color_ro(val):
            if isinstance(val, str) and val.endswith('%'):
                try:
                    ro_numeric = float(val.strip('%'))
                    if ro_numeric >= 100:
                        return 'background-color: #C6EFCE; color: #006100; font-weight: bold;'
                    elif 80 <= ro_numeric <= 99:
                        return 'background-color: #FFEB9C; color: #9C6500; font-weight: bold;'
                    elif ro_numeric < 80:
                        return 'background-color: #FFC7CE; color: #9C0006; font-weight: bold;'
                except:
                    pass
            return ''

        styled_df = df_pvt_summary_display.style.applymap(color_ro, subset=['R/O'])
        st.dataframe(styled_df, use_container_width=True)
    else:
        st.warning("Aucune donnée disponible pour le résumé par PVT")

    st.markdown("<br>", unsafe_allow_html=True)

    # 📊 TABLEAU 3 : Détails par VTO
    st.markdown(f'<div class="section-title">👥 DÉTAIL PAR VENDEUR - {mois_nom.upper()} {annee}</div>', unsafe_allow_html=True)

    if not df_reporting.empty:
        total_realisation_vto = df_reporting['REALISATION'].sum()
        df_reporting_display = df_reporting.copy()
        df_reporting_display.loc[len(df_reporting_display)] = ['', '', '', '', 'TOTAL', total_realisation_vto]

        st.dataframe(df_reporting_display, use_container_width=True)
    else:
        st.warning("Aucune donnée disponible pour le détail par VTO")

    st.markdown("<br>", unsafe_allow_html=True)

    # 📊 GRAPHIQUES
    st.markdown('<div class="section-title">📈 VISUALISATION DES PERFORMANCES</div>', unsafe_allow_html=True)

    if 'DR' in df_filtre.columns and not df_filtre.empty:
        df_drv = df_filtre.groupby('DR').size().reset_index(name='REALISATION')
        df_drv = df_drv.sort_values('REALISATION', ascending=False)

        total_ventes = df_drv['REALISATION'].sum()
        df_drv['POURCENTAGE'] = (df_drv['REALISATION'] / total_ventes * 100).round(1)

        col_dr1, col_dr2 = st.columns(2)

        with col_dr1:
            fig_pie_dr = px.pie(
                df_drv,
                values='REALISATION',
                names='DR',
                title=f'Distribution des ventes par DR - {mois_nom} {annee}',
                color_discrete_sequence=['#FF7900', '#FF5000', '#FF3000', '#E57200', '#CC6600', '#B35900'],
                hole=0.3
            )

            df_drv['LABEL'] = df_drv.apply(lambda row: f"{row['DR']}<br>{row['REALISATION']:,} ventes<br>({row['POURCENTAGE']}%)", axis=1)

            fig_pie_dr.update_traces(
                textposition='inside',
                textinfo='text',
                text=df_drv['LABEL'],
                textfont_size=10,
                marker=dict(line=dict(color='white', width=2)),
                hovertemplate="<b>%{label}</b><br>" +
                             "Réalisation: %{value:,}<br>" +
                             "Pourcentage: %{percent:.1%}<br>" +
                             "<extra></extra>"
            )

            fig_pie_dr.update_layout(
                title=dict(
                    text=f"Distribution des ventes par DR",
                    font=dict(size=16, family='Poppins', color='#FF7900'),
                    x=0.5,
                    y=0.95
                ),
                font=dict(family='Poppins', size=10),
                height=400,
                margin=dict(t=70, b=30, l=30, r=30),
                showlegend=False,
                hoverlabel=dict(
                    bgcolor="white",
                    font_size=11,
                    font_family="Poppins"
                )
            )

            st.plotly_chart(fig_pie_dr, use_container_width=True)

        with col_dr2:
            fig_bar_dr = go.Figure()

            df_drv_bar = df_drv.sort_values('REALISATION', ascending=False)

            fig_bar_dr.add_trace(go.Bar(
                x=df_drv_bar['DR'],
                y=df_drv_bar['REALISATION'],
                marker_color='#FF7900',
                text=df_drv_bar.apply(lambda row: f"{row['REALISATION']:,}<br>({row['POURCENTAGE']}%)", axis=1),
                textposition='outside',
                textfont=dict(size=10, color='#333'),
                textangle=0,
                hovertemplate="<b>%{x}</b><br>" +
                             "Réalisation: %{y:,}<br>" +
                             "Pourcentage: %{customdata}%<br>" +
                             "<extra></extra>",
                customdata=df_drv_bar['POURCENTAGE']
            ))

            fig_bar_dr.update_layout(
                title=dict(
                    text=f"Performances par DR - {mois_nom} {annee}",
                    font=dict(size=16, family='Poppins', color='#FF7900'),
                    x=0.5,
                    y=0.95
                ),
                xaxis_title='',
                yaxis_title='Réalisation',
                template='plotly_white',
                height=400,
                margin=dict(t=70, b=30, l=60, r=30),
                font=dict(family='Poppins', size=10),
                hoverlabel=dict(
                    bgcolor="white",
                    font_size=11,
                    font_family="Poppins"
                ),
                xaxis=dict(
                    tickangle=0,
                    tickfont=dict(size=10),
                    automargin=True,
                    showgrid=False,
                    categoryorder='total descending'
                ),
                yaxis=dict(
                    showgrid=True,
                    gridcolor='lightgray',
                    tickformat=','
                )
            )

            st.plotly_chart(fig_bar_dr, use_container_width=True)

    col_pvt, col_vto = st.columns(2)

    with col_pvt:
        if 'PVT' in df_filtre.columns and 'DR' in df_filtre.columns and not df_filtre.empty:
            df_pvt_summary_chart = df_filtre.groupby(['DR', 'PVT']).size().reset_index(name='REALISATION')
            df_top_pvt = df_pvt_summary_chart.nlargest(5, 'REALISATION')

            df_top_pvt = df_top_pvt.sort_values('REALISATION', ascending=True)

            fig_pvt = go.Figure()

            fig_pvt.add_trace(go.Bar(
                y=df_top_pvt['PVT'].str[:30] + '...',
                x=df_top_pvt['REALISATION'],
                orientation='h',
                marker_color='#009CA6',
                text=df_top_pvt['REALISATION'],
                textposition='outside',
                textfont=dict(size=10, color='#333'),
                hovertemplate="<b>%{y}</b><br>" +
                             "Réalisation: %{x:,}<br>" +
                             "DR: %{customdata}<br>" +
                             "<extra></extra>",
                customdata=df_top_pvt['DR']
            ))

            fig_pvt.update_layout(
                title=dict(
                    text=f"Top 5 PVT - {mois_nom} {annee}",
                    font=dict(size=16, family='Poppins', color='#009CA6'),
                    x=0.5,
                    y=0.95
                ),
                xaxis_title='Réalisation',
                yaxis_title='',
                template='plotly_white',
                height=350,
                margin=dict(t=70, b=30, l=200, r=30),
                font=dict(family='Poppins', size=10),
                hoverlabel=dict(
                    bgcolor="white",
                    font_size=11,
                    font_family="Poppins"
                ),
                yaxis=dict(
                    tickfont=dict(size=9),
                    automargin=True,
                    showgrid=False
                ),
                xaxis=dict(
                    showgrid=True,
                    gridcolor='lightgray',
                    tickformat=','
                )
            )

            st.plotly_chart(fig_pvt, use_container_width=True)

    with col_vto:
        if not df_reporting.empty:
            df_top10 = df_reporting.nlargest(10, 'REALISATION').copy()
            df_top10 = df_top10.sort_values('REALISATION', ascending=False)

            fig_top10 = go.Figure()

            fig_top10.add_trace(go.Bar(
                x=df_top10.apply(lambda row: f"{row['PRENOM_VENDEUR'][:10]}...", axis=1),
                y=df_top10['REALISATION'],
                marker_color='#FF5000',
                text=df_top10['REALISATION'],
                textposition='outside',
                textfont=dict(size=10, color='#333'),
                marker_line=dict(color='white', width=1),
                hovertemplate="<b>%{customdata[0]} %{customdata[1]}</b><br>" +
                             "Réalisation: %{y:,}<br>" +
                             "PVT: %{customdata[2]}<br>" +
                             "DR: %{customdata[3]}<br>" +
                             "<extra></extra>",
                customdata=df_top10[['PRENOM_VENDEUR', 'NOM_VENDEUR', 'PVT', 'DR']]
            ))

            fig_top10.update_layout(
                title=dict(
                    text=f"Top 10 Vendeurs - {mois_nom} {annee}",
                    font=dict(size=16, family='Poppins', color='#FF5000'),
                    x=0.5,
                    y=0.95
                ),
                xaxis_title='',
                yaxis_title='Réalisation',
                template='plotly_white',
                height=400,
                margin=dict(t=70, b=100, l=60, r=30),
                font=dict(family='Poppins', size=10),
                hoverlabel=dict(
                    bgcolor="white",
                    font_size=11,
                    font_family="Poppins"
                ),
                xaxis=dict(
                    tickangle=-45,
                    tickfont=dict(size=9),
                    automargin=True,
                    showgrid=False
                ),
                yaxis=dict(
                    showgrid=True,
                    gridcolor='lightgray',
                    tickformat=','
                )
            )

            st.plotly_chart(fig_top10, use_container_width=True)

# ====================
# INTERFACE PRINCIPALE
# ====================
def main():
    st.markdown('<div class="section-title">📅 SÉLECTION DE LA PÉRIODE</div>', unsafe_allow_html=True)

    current_date = datetime.now()
    current_year = current_date.year
    current_month = current_date.month

    col_month, col_year = st.columns(2)

    with col_month:
        selected_month = st.selectbox(
            "Mois",
            options=list(range(1, 13)),
            format_func=lambda x: calendar.month_name[x],
            index=current_month - 1
        )
        mois_nom = calendar.month_name[selected_month]

    with col_year:
        selected_year = st.selectbox(
            "Année",
            options=list(range(current_year - 2, current_year + 1)),
            index=2
        )

    st.markdown("<br>", unsafe_allow_html=True)

    st.markdown('<div class="section-title">📁 IMPORTATION DES DONNÉES</div>', unsafe_allow_html=True)

    uploaded_file = st.file_uploader(
        f"Importer le fichier Excel/CSV pour {mois_nom} {selected_year}",
        type=["xlsx", "csv", "xls"],
        help="Le fichier doit contenir les colonnes: MSISDN, ACCUEIL_VENDEUR, LOGIN_VENDEUR, AGENCE_VENDEUR, NOM_VENDEUR, PRENOM_VENDEUR, ETAT_IDENTIFICATION"
    )

    if uploaded_file:
        st.markdown("<br>", unsafe_allow_html=True)

        try:
            if uploaded_file.name.endswith('.csv'):
                df = pd.read_csv(uploaded_file, encoding='utf-8', sep=';')
            elif uploaded_file.name.endswith('.xls'):
                df = pd.read_excel(uploaded_file, engine='xlrd')
            else:
                xls = pd.ExcelFile(uploaded_file)
                sheet_names = xls.sheet_names
                if len(sheet_names) == 1:
                    selected_sheet = sheet_names[0]
                else:
                    selected_sheet = st.selectbox(
                        "🗂 Choisir la feuille à exploiter :",
                        options=sheet_names
                    )
                df = pd.read_excel(uploaded_file, sheet_name=selected_sheet)

            st.success(f"✅ Fichier chargé avec succès ! {len(df)} lignes trouvées")

            vto_df = load_vto()
            details = ["En Cours-Identification", "Identifie", "Identifie Photo"]

            objectif_mensuel = 960

            df_filtre, df_pvt_summary, df_dr_summary, df_reporting = process_monthly_data(
                df, vto_df, details, objectif_mensuel
            )

            if not df_filtre.empty:
                st.success(f"✅ Données traitées avec succès ! {len(df_filtre)} ventes mensuelles analysées")

                display_monthly_metrics(df_filtre, df_reporting, df_pvt_summary, df_dr_summary, mois_nom, selected_year)

                st.markdown('<div class="section-title">📥 TÉLÉCHARGEMENT DU RAPPORT</div>', unsafe_allow_html=True)

                try:
                    excel_buffer = generate_monthly_excel_report(
                        df_pvt_summary, df_dr_summary, df_reporting, mois_nom, selected_year
                    )

                    file_name = f"Reporting_Mensuel_{mois_nom}_{selected_year}.xlsx"

                    st.download_button(
                        label=f"📥 Télécharger le Reporting Mensuel complet",
                        data=excel_buffer,
                        file_name=file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_monthly_report",
                        use_container_width=True
                    )
                except Exception as e:
                    st.error(f"❌ Erreur lors de la génération du fichier Excel : {str(e)}")
            else:
                st.warning("⚠️ Aucune donnée filtrée disponible. Vérifiez votre fichier source et le référentiel VTO.")

        except Exception as e:
            st.error(f"❌ Erreur lors du traitement des données : {str(e)}")

# ====================
# EXÉCUTION PRINCIPALE
# ====================
if __name__ == "__main__":
    main()